"""
safetec_core – Combined Fleet + Driver seed script
====================================================
Imports truck/trailer data from the two fleet Excel files and driver data
from the drivers Excel file into your local development database.

Usage (run from the backend/ directory):
    python scripts/seed_all.py

Or to run only one section:
    python scripts/seed_all.py --fleet
    python scripts/seed_all.py --drivers

Requirements:
    pip install openpyxl
    Place the three Excel files in backend/scripts/seed_data/:
        FLEET_LIST_OBHI_TRUCKS_AND_TRAILORS_UPDATED.xlsx
        FLEET_LIST_OF_SAFETEC_TRUCKS_AND_TRAILORS_UPDATED.xlsx
        USE_THIS_Drivers_JAN_2026_updated.xlsx

The script is fully IDEMPOTENT – re-running it safely skips records that
already exist (matched by registration for trucks, employee number for
permanent drivers, and full name for casuals).
"""

import os
import sys
import argparse

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from sqlalchemy.orm import Session
from app.db.database import engine, SessionLocal
from app.models.models import (
    Base, BusinessEntity,
    Truck, Trailer, TruckStatus, TrailerStatus,
    Driver, DriverType,
)

# ── File paths ────────────────────────────────────────────────────────────────

DATA_DIR    = os.path.join(os.path.dirname(__file__), "seed_data")
OBHI_FILE   = os.path.join(DATA_DIR, "FLEET_LIST_OBHI_TRUCKS_AND_TRAILORS_UPDATED.xlsx")
SFT_FILE    = os.path.join(DATA_DIR, "FLEET_LIST_OF_SAFETEC_TRUCKS_AND_TRAILORS_UPDATED.xlsx")
DRIVER_FILE = os.path.join(DATA_DIR, "USE_THIS_Drivers_JAN_2026_updated.xlsx")

# ── Entity codes – adjust if your DB uses different codes ─────────────────────
OBHI_CODE = "OBHI"
SFT_CODE  = "SFT"


# ─────────────────────────────────────────────────────────────────────────────
# SHARED HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def get_entity(db: Session, code: str) -> BusinessEntity:
    entity = db.query(BusinessEntity).filter(BusinessEntity.code == code).first()
    if not entity:
        raise SystemExit(
            f"\n✗  Entity with code '{code}' not found in the database.\n"
            f"    Make sure the entity exists before running this seed."
        )
    return entity


def clean_reg(value) -> str | None:
    """Normalise a registration plate: strip, uppercase, take first if dual."""
    if not value:
        return None
    s = str(value).strip().upper()
    if "/" in s:
        s = s.split("/")[0].strip()
    return s or None


def parse_make_model(raw: str):
    """'SCANIA G460' → ('SCANIA', 'G460').  'SCANIA ' → ('SCANIA', None)."""
    parts = str(raw).strip().split(None, 1)
    make  = parts[0].upper() if parts else ""
    model = parts[1].strip() if len(parts) > 1 else None
    return make, model


def clean_str(v) -> str | None:
    if not v:
        return None
    s = str(v).strip()
    return s or None


# ─────────────────────────────────────────────────────────────────────────────
# FLEET SEED
# ─────────────────────────────────────────────────────────────────────────────

def truck_exists(db: Session, entity_id: int, registration: str) -> bool:
    return db.query(Truck).filter(
        Truck.entity_id == entity_id,
        Truck.registration == registration,
    ).first() is not None


def import_obhi_fleet(db: Session, entity: BusinessEntity):
    """
    OBHI fleet Excel columns (0-indexed):
      0=make, 1=truck_reg, 2=truck_vin,
      3=label_t1, 4=trailer1_reg, 5=trailer1_vin,
      6=label_t2, 7=trailer2_reg, 8=trailer2_vin,
      9=maintenance_supplier (ignored)
    """
    from openpyxl import load_workbook
    wb = load_workbook(OBHI_FILE, read_only=True)
    ws = wb.active
    created = skipped = 0

    for row in ws.iter_rows(values_only=True):
        raw_make = row[0]
        if not raw_make or str(raw_make).strip().upper() in ("OBHI PTY LTD", ""):
            continue

        truck_reg = clean_reg(row[1])
        if not truck_reg:
            continue

        if truck_exists(db, entity.id, truck_reg):
            print(f"    ↷   {truck_reg} – already exists, skipping")
            skipped += 1
            continue

        make, model = parse_make_model(raw_make)
        truck = Truck(
            entity_id=entity.id,
            make=make,
            model=model,
            registration=truck_reg,
            vin=clean_str(row[2]),
            status=TruckStatus.active,
        )
        db.add(truck)
        db.flush()

        for slot, reg_col, vin_col in [(1, 4, 5), (2, 7, 8)]:
            t_reg = clean_reg(row[reg_col]) if len(row) > reg_col else None
            t_vin = clean_str(row[vin_col]) if len(row) > vin_col else None
            if t_reg:
                db.add(Trailer(
                    truck_id=truck.id,
                    entity_id=entity.id,
                    slot=slot,
                    registration=t_reg,
                    vin=t_vin,
                    status=TrailerStatus.active,
                ))

        created += 1

    db.commit()
    print(f"    ✓  OBHI fleet: {created} trucks created, {skipped} skipped")


def import_sft_fleet(db: Session, entity: BusinessEntity):
    """
    Safetec fleet Excel columns (0-indexed):
      0=fleet_num, 1=make, 2=truck_reg, 3=driver_name,
      4=label_t1, 5=trailer1_reg, 6=label_t2, 7=trailer2_reg
    """
    from openpyxl import load_workbook
    wb = load_workbook(SFT_FILE, read_only=True)
    ws = wb.active
    created = skipped = 0

    for row in ws.iter_rows(values_only=True):
        fleet_num = row[0]
        raw_make  = row[1]
        if not fleet_num or not raw_make:
            continue
        if not str(fleet_num).strip().isdigit():
            continue

        truck_reg = clean_reg(row[2])
        if not truck_reg:
            continue

        if truck_exists(db, entity.id, truck_reg):
            print(f"    ↷   {truck_reg} – already exists, skipping")
            skipped += 1
            continue

        make, model = parse_make_model(raw_make)
        truck = Truck(
            entity_id=entity.id,
            fleet_number=str(fleet_num).strip(),
            make=make,
            model=model,
            registration=truck_reg,
            driver_name=clean_str(row[3]),
            status=TruckStatus.active,
        )
        db.add(truck)
        db.flush()

        for slot, reg_col in [(1, 5), (2, 7)]:
            t_reg = clean_reg(row[reg_col]) if len(row) > reg_col else None
            if t_reg:
                db.add(Trailer(
                    truck_id=truck.id,
                    entity_id=entity.id,
                    slot=slot,
                    registration=t_reg,
                    status=TrailerStatus.active,
                ))

        created += 1

    db.commit()
    print(f"    ✓  Safetec fleet: {created} trucks created, {skipped} skipped")


def seed_fleet(db: Session):
    print("\n🚛  Seeding fleet...\n")

    obhi = get_entity(db, OBHI_CODE)
    sft  = get_entity(db, SFT_CODE)

    print(f"  → OBHI fleet  (entity: {obhi.name})")
    import_obhi_fleet(db, obhi)

    print(f"\n  → Safetec fleet  (entity: {sft.name})")
    import_sft_fleet(db, sft)


# ─────────────────────────────────────────────────────────────────────────────
# DRIVER SEED
# ─────────────────────────────────────────────────────────────────────────────

def driver_exists_by_emp(db: Session, entity_id: int, emp_no: str) -> bool:
    return db.query(Driver).filter(
        Driver.entity_id == entity_id,
        Driver.employee_number == emp_no,
    ).first() is not None


def driver_exists_by_name(db: Session, entity_id: int, first: str, last: str) -> bool:
    return db.query(Driver).filter(
        Driver.entity_id == entity_id,
        Driver.first_name == first,
        Driver.last_name == last,
    ).first() is not None


def find_truck(db: Session, entity_id: int, registration: str) -> Truck | None:
    """Look up a truck by registration, scoped to the entity."""
    if not registration:
        return None
    reg = registration.strip().upper()
    return db.query(Truck).filter(
        Truck.entity_id == entity_id,
        Truck.registration == reg,
    ).first()


def seed_drivers(db: Session):
    from openpyxl import load_workbook

    print("\n👤  Seeding drivers...\n")

    sft  = get_entity(db, SFT_CODE)
    obhi = get_entity(db, OBHI_CODE)

    def truck_id(entity: BusinessEntity, reg: str) -> int | None:
        t = find_truck(db, entity.id, reg)
        return t.id if t else None

    # ── Permanent drivers (all belong to SFT entity) ──────────────────────
    permanent_drivers = [
        ("THEM002",  "Luizi",               "Kambinda",   None,         None,              None,           None,   None),
        ("THEM004",  "Ndondi Simon",        "Ndlala",     "KMR411EC",   None,              None,           None,   None),
        ("THEM008",  "Abongile",            "Mana",       "KNS946EC",   None,              None,           None,   None),
        ("THEM011",  "Kliffet",             "Williams",   "KBD393EC",   None,              None,           None,   None),
        ("THEM-021", "Resty",               "Hollander",  "JZS095EC",   None,              None,           None,   None),
        ("THEM-026", "Kgosiema Roy",        "Setlhare",   "KPH748EC",   None,              None,           None,   None),
        ("THEM-027", "Sibusiso Arnold",     "Silomo",     "KNF472EC",   None,              None,           None,   None),
        ("THEM028",  "Sam",                 "Mahlamba",   "KMT481EC",   "8811015438080",   "2535529164",   None,   None),
        ("THEM031",  "Petrus Jacobus",      "Pretorius",  "KRM473EC",   "6312265270085",   None,           None,   None),
        ("THEM033",  "Andries Moitsemang",  "Mathe",      "KGJ578EC",   None,              None,           None,   None),
        ("THEM037",  "Itemogeng Job",       "Dimengu",    "KLC159EC",   None,              None,           None,   None),
        ("THEM043",  "Stoffel",             "Sefadi",     "KSZ397EC",   None,              None,           "FNB",  "62418950045"),
        ("THEM045",  "Samuel",              "Domingo",    "KDZ484EC",   None,              None,           None,   None),
        ("THEM046",  "Phikane",             "Matiwane",   "KFK772EC",   None,              None,           None,   None),
        ("THEM047",  "Kgololo Joseph",      "Mocwane",    "KRC830EC",   None,              None,           None,   None),
        ("THEM048",  "Randall",             "Bosch",      "KDJ034EC",   None,              None,           None,   None),
        ("THEM049",  "Ruwatira",            "Tapiwa",     "KNP861EC",   None,              None,           None,   None),
        ("THEM050",  "Samuel",              "Mmotsi",     "KKF401EC",   None,              None,           None,   None),
        ("THEM051",  "Freek",               "Boss",       "JZG083EC",   None,              None,           None,   None),
        ("THEM054",  "Letlhogonolo Aubry",  "Gaebidiwe",  "KRW188EC",   None,              None,           None,   None),
        ("THEM056",  "Pieter",              "Hendrikcs",  "JYC247EC",   None,              None,           None,   None),
        ("THEM057",  "Miguel",              "Maseka",     "KFT767EC",   None,              None,           None,   None),
        ("THEM058",  "Simon Tshidiso",      "Sethunya",   "JXP816EC",   None,              None,           None,   None),
        ("THEM060",  "Johny",               "Jacobs",     "JXP816EC",   None,              None,           None,   None),
        ("THEM061",  "David",               "Mohapeloa",  "KDZ484EC",   None,              None,           None,   None),
        ("THEM059",  "Aviwe",               "Hoza",       None,         None,              None,           None,   None),
        ("THEM052",  "Ayabulela",           "Mdingi",     None,         None,              None,           None,   None),
    ]

    perm_created = perm_skipped = 0
    for emp_no, first, last, truck_reg, id_no, tax_no, bank_name, bank_acc in permanent_drivers:
        if driver_exists_by_emp(db, sft.id, emp_no):
            print(f"    ↷   {emp_no} {first} {last} – already exists, skipping")
            perm_skipped += 1
            continue

        t_id = truck_id(sft, truck_reg) if truck_reg else None

        driver = Driver(
            entity_id=sft.id,
            truck_id=t_id,
            employee_number=emp_no,
            first_name=first,
            last_name=last,
            driver_type=DriverType.permanent,
            id_number=id_no,
            tax_number=tax_no,
            bank_name=bank_name,
            bank_account_number=bank_acc,
            is_active=True,
        )
        db.add(driver)
        perm_created += 1

    db.flush()
    db.commit()
    print(f"    ✓  Permanent drivers: {perm_created} created, {perm_skipped} skipped")

    # ── Casual drivers ────────────────────────────────────────────────────
    casual_drivers = [
        (sft,  "Gerald",  "Casual",  "KCY733EC",  None,       None,           "Casual driver – Safetec"),
        (sft,  "Maskito", "Casual",  "KRM473EC",  None,       None,           "Casual driver – shares truck with Petrus"),
        (sft,  "Tumiso",  "Casual",  "JYC247EC",  None,       None,           "Casual driver – Safetec"),
        (sft,  "Mathews", "Casual",  "JRC117EC",  None,       None,           "Casual driver – Safetec"),
        (sft,  "Adriaan", "Casual",  "KPS102EC",  None,       None,           "Casual driver – Safetec"),
        (sft,  "Garth",   "Casual",  "KDJ034EC",  None,       None,           "Casual driver – Safetec"),
        (obhi, "Jerry",   "Bandeys", "KPS629EC",  "CAPITEC",  "1169321362",   "Casual driver – OBHI (Bandeys)"),
        (obhi, "Zamani",  "Bandeys", "KSC007EC",  "CAPITEC",  "1267731808",   "Casual driver – OBHI (Bandeys)"),
        (obhi, "Amos",    "Bandeys", "EKV375GP",  "CAPITEC",  "1323790177",   "Casual driver – OBHI (Bandeys)"),
    ]

    cas_created = cas_skipped = 0
    for entity, first, last, truck_reg, bank_name, bank_acc, notes in casual_drivers:
        if driver_exists_by_name(db, entity.id, first, last):
            print(f"    ↷   {first} {last} ({entity.code}) – already exists, skipping")
            cas_skipped += 1
            continue

        t_id = truck_id(entity, truck_reg)
        if not t_id and truck_reg:
            other = sft if entity == obhi else obhi
            t_id = truck_id(other, truck_reg)

        driver = Driver(
            entity_id=entity.id,
            truck_id=t_id,
            first_name=first,
            last_name=last,
            driver_type=DriverType.casual,
            bank_name=bank_name,
            bank_account_number=bank_acc,
            notes=notes,
            is_active=True,
        )
        db.add(driver)
        cas_created += 1

    db.commit()
    print(f"    ✓  Casual drivers:    {cas_created} created, {cas_skipped} skipped")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Seed fleet and/or drivers into safetec_core.")
    parser.add_argument("--fleet",   action="store_true", help="Seed fleet only")
    parser.add_argument("--drivers", action="store_true", help="Seed drivers only")
    args = parser.parse_args()

    run_fleet   = args.fleet   or (not args.fleet and not args.drivers)
    run_drivers = args.drivers or (not args.fleet and not args.drivers)

    print("=" * 60)
    print("  safetec_core – Seed Script")
    print("=" * 60)

    Base.metadata.create_all(bind=engine)

    db: Session = SessionLocal()
    try:
        if run_fleet:
            seed_fleet(db)
        if run_drivers:
            seed_drivers(db)

        print("\n" + "=" * 60)
        print("  ✓  Seed complete.")
        print("=" * 60 + "\n")

    except SystemExit:
        raise
    except Exception as e:
        db.rollback()
        print(f"\n✗  Seed failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()
