"""Supplier Summary Excel export (openpyxl).

One worksheet per supplier — the way the manual supplier workbook is kept —
with the year's invoices grouped under a banner per month, Paid / Outstanding /
Total under each month, and a period total at the bottom. A
Summary sheet up front holds the supplier × month grid (incl VAT) with Paid /
Outstanding / Total rows beneath it.

Rows are exactly what the on-screen Supplier Summary shows
(reports._build_month_detail: invoice-date basis with Manage→Move pins,
report-only reclassifications, user exclusions dropped), so every figure ties
back to the report and to Income vs Expenses.
"""

import io
import re

from app.core.constants import MONTH_NAMES_0 as MONTH_NAMES

# Characters Excel forbids in sheet names.
_SHEET_BAD = re.compile(r"[\[\]:*?/\\]")


def _sheet_title(name: str, used: set) -> str:
    clean = _SHEET_BAD.sub(" ", name).strip()[:31] or "Supplier"
    base, n = clean, 2
    while clean.upper() in used:
        suffix = f" ({n})"
        clean = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(clean.upper())
    return clean


# Per-block sums: excl / vat / incl overall, and the incl-VAT split into paid
# vs outstanding (Suppliers-overview rule: is_paid flag set → paid, else the
# whole amount is outstanding).
_SUM_KEYS = ("excl", "vat", "incl", "paid_excl", "paid_vat", "paid_incl",
             "out_excl", "out_vat", "out_incl")


def _empty_sums() -> dict:
    return {k: 0.0 for k in _SUM_KEYS}


def _accumulate(sums: dict, r: dict) -> None:
    pre = "paid_" if r.get("is_paid") else "out_"
    for k, field in (("excl", "amount_excl"), ("vat", "vat"), ("incl", "amount_incl")):
        sums[k] += r[field]
        sums[pre + k] += r[field]


def generate_supplier_summary_excel(entity, period_label: str, month_rows: dict) -> bytes:
    """month_rows: {(year, month): [row, ...]} — each month's non-excluded
    supplier rows from _build_month_detail (date / invoice_number / vehicle_reg
    / description / amount_excl / vat / amount_incl / vat_applicable).
    period_label titles the workbook (e.g. "2026", "August 2026",
    "15 Jun 2026 – 20 Aug 2026")."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Group per supplier, months kept in calendar order.
    suppliers: dict[str, dict] = {}
    for ym in sorted(month_rows):
        for r in month_rows[ym]:
            name = r["supplier_name"] or "General Expenses"
            sup = suppliers.setdefault(name.upper(), {"name": name, "months": {}})
            sup["months"].setdefault(ym, []).append(r)
    ordered = sorted(suppliers.values(), key=lambda s: s["name"].upper())
    months_present = sorted(month_rows)

    # Month headings only carry the year when the export spans more than one.
    multi_year = len({y for y, _ in months_present}) > 1

    def month_label(ym):
        y, m = ym
        return MONTH_NAMES[m - 1].upper() + (f" {y}" if multi_year else "")

    wb = Workbook()

    NUM = "#,##0.00"
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    HDR_FILL = PatternFill("solid", fgColor="4B5563")
    MONTH_FILL = PatternFill("solid", fgColor="0D9488")  # teal banner, like the manual sheet
    TOTAL_FILL = PatternFill("solid", fgColor="E5E7EB")
    RIGHT = Alignment(horizontal="right", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def amount(ws, row, col, value, bold=False, fill=None):
        c = ws.cell(row=row, column=col, value=round(value, 2) if value is not None else None)
        c.number_format = NUM
        c.alignment = RIGHT
        c.border = border
        c.font = Font(bold=bold, size=9)
        if fill:
            c.fill = fill
        return c

    def label(ws, row, col, value, bold=False, fill=None, color="111111", size=9):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=color, size=size)
        c.alignment = LEFT
        c.border = border
        if fill:
            c.fill = fill
        return c

    def totals_block(ws, row, title, sums, fill, color, merge_label=False, split=True):
        """PAID / OUTSTANDING / TOTAL rows for one month — or, with
        split=False, just the TOTAL row (the period total at the foot of a
        multi-month tab). Paid and Outstanding sit on the light total fill; the
        TOTAL row takes the block's own fill/colour. Returns the row after the
        block."""
        def line(r, text, pre, row_fill, row_color):
            label(ws, r, 1, text, bold=True, fill=row_fill, color=row_color)
            label(ws, r, 2, "", fill=row_fill)
            if merge_label:
                # Merged across A:B so long period labels ("15 Jun 2026 – …")
                # aren't clipped by the narrow date column.
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            for col, k in ((3, "excl"), (4, "vat"), (5, "incl")):
                c = amount(ws, r, col, sums[pre + k], bold=True, fill=row_fill)
                c.font = Font(bold=True, color=row_color, size=9)
            label(ws, r, 6, "", fill=row_fill)
            label(ws, r, 7, "", fill=row_fill)

        if split:
            line(row, f"{title} PAID", "paid_", TOTAL_FILL, "111111")
            line(row + 1, f"{title} OUTSTANDING", "out_", TOTAL_FILL, "111111")
            row += 2
        line(row, f"{title} TOTAL", "", fill, color)
        return row + 1

    ent_label = f"{getattr(entity, 'code', '') or ''} — {getattr(entity, 'name', '') or ''}".strip(" —")

    # ── Summary sheet: supplier × month grid (incl VAT) ───────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 40
    for i in range(2, len(months_present) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 16 if multi_year else 14

    ws.cell(row=1, column=1, value=f"Supplier Summary — {period_label}").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=ent_label).font = Font(color="6B7280", size=10)
    ws.cell(row=3, column=1,
            value="Incl-VAT totals per supplier per month. Each supplier's invoices are on its own tab.").font = \
        Font(color="6B7280", size=9)

    row = 5
    label(ws, row, 1, "Supplier", bold=True, fill=HDR_FILL, color="FFFFFF")
    for i, m in enumerate(months_present):
        c = label(ws, row, 2 + i, month_label(m), bold=True, fill=HDR_FILL, color="FFFFFF")
        c.alignment = RIGHT
    c = label(ws, row, 2 + len(months_present), "TOTAL", bold=True, fill=HDR_FILL, color="FFFFFF")
    c.alignment = RIGHT
    row += 1

    # Paid / Outstanding follow the Suppliers overview rule: an invoice is paid
    # when its is_paid flag is set, otherwise the full amount is outstanding.
    paid_by_month = {m: 0.0 for m in months_present}
    outst_by_month = {m: 0.0 for m in months_present}
    for sup in ordered:
        label(ws, row, 1, sup["name"])
        sup_total = 0.0
        for i, m in enumerate(months_present):
            rows = sup["months"].get(m, [])
            mt = sum(r["amount_incl"] for r in rows)
            paid_by_month[m] += sum(r["amount_incl"] for r in rows if r.get("is_paid"))
            outst_by_month[m] += sum(r["amount_incl"] for r in rows if not r.get("is_paid"))
            sup_total += mt
            amount(ws, row, 2 + i, mt if rows else None)
        amount(ws, row, 2 + len(months_present), sup_total, bold=True)
        row += 1

    for title, by_month in (("PAID", paid_by_month), ("OUTSTANDING", outst_by_month)):
        label(ws, row, 1, title, bold=True, fill=TOTAL_FILL)
        for i, m in enumerate(months_present):
            amount(ws, row, 2 + i, by_month[m], bold=True, fill=TOTAL_FILL)
        amount(ws, row, 2 + len(months_present), sum(by_month.values()), bold=True, fill=TOTAL_FILL)
        row += 1
    label(ws, row, 1, "TOTAL", bold=True, fill=HDR_FILL, color="FFFFFF")
    for i, m in enumerate(months_present):
        amount(ws, row, 2 + i, paid_by_month[m] + outst_by_month[m], bold=True, fill=HDR_FILL) \
            .font = Font(bold=True, color="FFFFFF", size=9)
    amount(ws, row, 2 + len(months_present),
           sum(paid_by_month.values()) + sum(outst_by_month.values()), bold=True, fill=HDR_FILL) \
        .font = Font(bold=True, color="FFFFFF", size=9)
    ws.freeze_panes = "B6"

    # ── One sheet per supplier ────────────────────────────────────────────────
    COLS = ["Date", "Invoice #", "Excl VAT", "VAT", "Incl VAT", "Vehicle Reg", "Description"]
    WIDTHS = [12, 20, 13, 12, 13, 14, 60]
    used_titles = {"SUMMARY"}

    for sup in ordered:
        ws = wb.create_sheet(_sheet_title(sup["name"], used_titles))
        for i, w in enumerate(WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.cell(row=1, column=1, value=sup["name"]).font = Font(bold=True, size=13)
        ws.cell(row=2, column=1, value=f"{ent_label} · {period_label}").font = Font(color="6B7280", size=10)

        row = 4
        grand = _empty_sums()
        for m in sorted(sup["months"]):
            rows = sup["months"][m]
            # Month banner across the full width, like the manual sheet's teal band.
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLS))
            label(ws, row, 1, month_label(m), bold=True,
                  fill=MONTH_FILL, color="FFFFFF", size=10)
            for col in range(2, len(COLS) + 1):
                ws.cell(row=row, column=col).fill = MONTH_FILL
                ws.cell(row=row, column=col).border = border
            row += 1

            for i, h in enumerate(COLS, start=1):
                c = label(ws, row, i, h, bold=True, fill=HDR_FILL, color="FFFFFF")
                if 3 <= i <= 5:
                    c.alignment = RIGHT
            row += 1

            mt = _empty_sums()
            for r in rows:
                label(ws, row, 1, r["date"] or "")
                inv_no = r["invoice_number"] or ""
                label(ws, row, 2, inv_no + ("" if r.get("vat_applicable", True) else "  (NON-VAT)"))
                amount(ws, row, 3, r["amount_excl"])
                amount(ws, row, 4, r["vat"])
                amount(ws, row, 5, r["amount_incl"])
                label(ws, row, 6, r.get("vehicle_reg") or "")
                label(ws, row, 7, r["description"] or "")
                _accumulate(mt, r)
                row += 1

            row = totals_block(ws, row, month_label(m), mt, TOTAL_FILL, "111111")
            for k in grand:
                grand[k] += mt[k]
            row += 1

        # With a single month on the tab this would just repeat the month total.
        if len(sup["months"]) > 1:
            totals_block(ws, row, period_label.upper(), grand, HDR_FILL, "FFFFFF", merge_label=True, split=False)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
