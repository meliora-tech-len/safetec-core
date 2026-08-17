"""
Costing export generators — PDF (reportlab) and Excel (openpyxl).
Both include the entity letterhead/logo as a header.
"""

import io
from decimal import Decimal
from datetime import date

from app.core.constants import MONTH_NAMES_1 as MONTH_NAMES

D0 = Decimal("0")


def _diesel_label(name) -> str:
    """"<Supplier> Diesel" label — only appends "Diesel" when the supplier name
    doesn't already end in it (e.g. "Intsimbi Diesel" stays as-is, "Oukop" →
    "Oukop Diesel"), so the word is never doubled up."""
    n = (name or "").strip()
    return n if n.upper().endswith("DIESEL") else f"{n} Diesel"


def _fmt(v) -> str:
    try:
        return f"R {float(v):,.2f}"
    except (TypeError, ValueError):
        return "—"


def _fmtd(v) -> str:
    if v is None:
        return ""
    if isinstance(v, date):
        return v.strftime("%d-%m-%Y")
    return str(v)[:10]


def _tonnes(v) -> str:
    try:
        return f"{float(v):.2f}"
    except (TypeError, ValueError):
        return "—"


def _load_label(l) -> str:
    """One-line description of a load for the collapsed Excel detail rows."""
    parts = [f"    {_fmtd(getattr(l, 'load_date', None))}"]
    mine = getattr(l, "mine_name", None)
    if mine:
        parts.append(str(mine))
    slip = getattr(l, "slip_number", None)
    if slip:
        parts.append(f"slip {slip}")
    tonnes = getattr(l, "tonnes", None)
    if tonnes is not None:
        tr = f"{_tonnes(tonnes)}t"
        rate = getattr(l, "subcontractor_rate", None)
        if rate is not None:
            tr += f" @ {_fmt(rate)}"
        parts.append(tr)
    return "  ·  ".join(parts)


# ── PDF ───────────────────────────────────────────────────────────────────────

def generate_costing_pdf(costing, entity) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
        Image as RLImage,
    )
    from app.services.pdf_generator import _load_logo, _FONT_NORMAL, _FONT_BOLD, _hex_to_color

    accent_hex = getattr(entity, "primary_color", None) or "#1a1a2e"
    accent   = _hex_to_color(accent_hex)
    black    = colors.HexColor("#111111")
    gray_mid = colors.HexColor("#6b7280")
    gray_lt  = colors.HexColor("#f3f4f6")
    white    = colors.white
    hdr_bg   = colors.HexColor("#4B5563")
    pos_col  = colors.HexColor("#15803d")
    neg_col  = colors.HexColor("#dc2626")
    border_c = colors.HexColor("#e5e7eb")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=16*mm, leftMargin=16*mm,
                            topMargin=10*mm, bottomMargin=14*mm)

    def sty(name, **kw):
        d = dict(fontName=_FONT_NORMAL, fontSize=9, textColor=black, leading=12)
        d.update(kw)
        return ParagraphStyle(name, **d)

    s_title   = sty("title", fontSize=13, fontName=_FONT_BOLD, leading=17)
    s_sub     = sty("sub",   fontSize=9,  textColor=gray_mid,  leading=13)
    s_hdr     = sty("hdr",   fontSize=8,  fontName=_FONT_BOLD, textColor=white,    alignment=TA_CENTER)
    s_hdr_r   = sty("hdr_r", fontSize=8,  fontName=_FONT_BOLD, textColor=white,    alignment=TA_RIGHT)
    s_lbl     = sty("lbl",   fontSize=8,  textColor=black)
    s_val     = sty("val",   fontSize=8,  textColor=black,     alignment=TA_RIGHT)
    s_bold    = sty("bold",  fontSize=8,  fontName=_FONT_BOLD, textColor=black)
    s_bold_r  = sty("br",    fontSize=8,  fontName=_FONT_BOLD, textColor=black,    alignment=TA_RIGHT)
    s_tot     = sty("tot",   fontSize=8,  fontName=_FONT_BOLD, textColor=white)
    s_tot_r   = sty("totr",  fontSize=8,  fontName=_FONT_BOLD, textColor=white,    alignment=TA_RIGHT)
    s_grp     = sty("grp",   fontSize=7.5, textColor=gray_mid, fontName=_FONT_BOLD)
    s_grp_r   = sty("grpr",  fontSize=7.5, textColor=gray_mid, fontName=_FONT_BOLD, alignment=TA_RIGHT)
    s_sec     = sty("sec",   fontSize=9.5, textColor=black,    fontName=_FONT_BOLD)
    s_dsup    = sty("dsup",  fontSize=8.5, textColor=black,    fontName=_FONT_BOLD)
    s_subhdr  = sty("subh",  fontSize=9.5, textColor=gray_mid, fontName=_FONT_BOLD)
    s_net_pos = sty("np",    fontSize=9,  fontName=_FONT_BOLD, textColor=pos_col,  alignment=TA_RIGHT)
    s_net_neg = sty("nn",    fontSize=9,  fontName=_FONT_BOLD, textColor=neg_col,  alignment=TA_RIGHT)
    s_smry    = sty("sm",    fontSize=9,  fontName=_FONT_BOLD, textColor=white,    alignment=TA_RIGHT)
    s_smry_l  = sty("sml",   fontSize=9,  fontName=_FONT_BOLD, textColor=white)
    s_dl      = sty("dl",    fontSize=7.5, textColor=gray_mid)
    s_dl_r    = sty("dlr",   fontSize=7.5, textColor=gray_mid, alignment=TA_RIGHT)

    story = []

    # ── Letterhead / logo ─────────────────────────────────────────────────────
    lh_url  = getattr(entity, "letterhead_url",  None)
    lh_path = getattr(entity, "letterhead_path", None)
    added_header = False
    if lh_url or lh_path:
        lh_bytes = _load_logo(lh_url, lh_path)
        if lh_bytes:
            try:
                img = RLImage(io.BytesIO(lh_bytes), width=174*mm, height=105*mm, kind="proportional")
                img.hAlign = "CENTER"
                story.append(img)
                story.append(Spacer(1, 4*mm))
                added_header = True
            except Exception:
                pass

    if not added_header:
        logo_url  = getattr(entity, "logo_url",  None)
        logo_path = getattr(entity, "logo_path", None)
        if logo_url or logo_path:
            lb = _load_logo(logo_url, logo_path)
            if lb:
                try:
                    img = RLImage(io.BytesIO(lb), width=55*mm, height=30*mm, kind="proportional")
                    img.hAlign = "LEFT"
                    story.append(img)
                    story.append(Spacer(1, 3*mm))
                    added_header = True
                except Exception:
                    pass
        if not added_header:
            ent_name = getattr(entity, "name", "") or ""
            story.append(Paragraph(ent_name, sty("en", fontSize=14, fontName=_FONT_BOLD)))
            story.append(Spacer(1, 3*mm))

    # ── Title ─────────────────────────────────────────────────────────────────
    period = f"{MONTH_NAMES[costing.month]} {costing.year}"
    sub_name = costing.subcontractor.name if costing.subcontractor else ""
    story.append(Paragraph(f"Subcontractor Costing — {sub_name}", s_title))
    story.append(Paragraph(f"Period: {period}", s_sub))
    story.append(Spacer(1, 5*mm))

    # ── Per-truck tables ──────────────────────────────────────────────────────
    col_w = [58*mm, 30*mm, 30*mm, 30*mm, 30*mm]  # label, income excl, income incl, exp incl, exp excl
    ts_hdr = TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0), hdr_bg),
        ("TEXTCOLOR",    (0, 0), (-1, 0), white),
        ("FONTNAME",     (0, 0), (-1, 0), _FONT_BOLD),
        ("FONTSIZE",     (0, 0), (-1, 0), 7.5),
        ("ALIGN",        (1, 0), (-1, 0), "RIGHT"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ("GRID",         (0, 0), (-1, -1), 0.4, border_c),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
    ])

    def P(txt, style): return Paragraph(str(txt), style)
    dash = P("—", s_val)

    for td in costing.trucks:
        t = td.truck
        reg = t.registration or ""
        fleet = f" · {t.fleet_number}" if t.fleet_number else ""
        story.append(Paragraph(f"Truck: {reg}{fleet}", sty("trk", fontSize=10, fontName=_FONT_BOLD, spaceBefore=4)))
        story.append(Spacer(1, 1*mm))

        inc_e = float(td.income_excl_vat or 0)
        inc_i = float(td.income_incl_vat or 0)
        loads_inc_e = float(td.loads_income_excl_vat or 0)
        loads_inc_i = float(td.loads_income_incl_vat or 0)
        exp_i = float(td.total_expenses_incl_vat or 0)
        exp_e = float(td.total_expenses_excl_vat or 0)
        net   = float(td.net_payable or 0)
        is_vat = costing.is_vat_registered

        rows = [
            [P("", s_hdr), P("Income Excl VAT", s_hdr_r), P("Income Incl VAT", s_hdr_r),
             P("Expenses Incl VAT", s_hdr_r), P("Expenses Excl VAT", s_hdr_r)],
            [P(f"Loads ({len(td.loads)})", s_sec),
             P(_fmt(loads_inc_e), s_val),
             P(_fmt(loads_inc_i), s_val) if is_vat else dash,
             dash, dash],
        ]
        for mi in (td.manual_incomes or []):
            m_excl = float(mi.amount_excl_vat or 0)
            m_incl = float(mi.amount_incl_vat or 0)
            rows.append([
                P(mi.description, s_lbl),
                P(_fmt(m_excl), s_val),
                P(_fmt(m_incl), s_val) if is_vat else dash,
                dash, dash,
            ])
        rows.append([P("Admin Fee", s_lbl), dash, dash, P(_fmt(td.admin_fee), s_val), dash])

        # Diesel groups
        for dg in (td.diesel_groups or []):
            rows.append([P(_diesel_label(dg.supplier_name), s_lbl), dash, dash, dash, P(_fmt(dg.tot_excl_admin_fee), s_val)])
            rows.append([P(f"{_diesel_label(dg.supplier_name)} Admin Fee", s_lbl), dash, dash, P(_fmt(dg.tot_admin_fee_incl), s_val), dash])

        # Supplier invoices
        if td.supplier_invoices:
            rows.append([P("Supplier Invoices", s_sec), P("", s_val), P("", s_val), P("", s_val), P("", s_val)])
            for inv in td.supplier_invoices:
                sup = inv.supplier_name or f"Supplier #{inv.supplier_id}"
                if inv.vat_applicable:
                    rows.append([P(sup, s_lbl), dash, dash, P(_fmt(inv.amount), s_val), dash])
                else:
                    rows.append([P(sup, s_lbl), dash, dash, dash, P(_fmt(inv.amount), s_val)])

        # Totals row (white text — this row gets a dark background below)
        net_s = s_net_pos if net >= 0 else s_net_neg
        dash_w = P("—", s_tot_r)
        rows.append([
            P("Totals", s_tot),
            P(_fmt(inc_e), s_tot_r),
            P(_fmt(inc_i), s_tot_r) if is_vat else dash_w,
            P(_fmt(exp_i), s_tot_r),
            P(_fmt(exp_e), s_tot_r),
        ])

        tbl = Table(rows, colWidths=col_w, repeatRows=1)
        tbl.setStyle(ts_hdr)
        # Alternate rows
        for i in range(2, len(rows)):
            if i % 2 == 0:
                tbl.setStyle(TableStyle([("BACKGROUND", (0, i), (-1, i), gray_lt)]))
        # Totals row style
        last = len(rows) - 1
        tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0, last), (-1, last), colors.HexColor("#374151")),
            ("TEXTCOLOR",   (0, last), (-1, last), white),
            ("FONTNAME",    (0, last), (-1, last), _FONT_BOLD),
        ]))
        story.append(tbl)

        # Net payable
        net_color = "#15803d" if net >= 0 else "#dc2626"
        story.append(Spacer(1, 1*mm))
        net_row = Table(
            [[P(f"Net Payable: {_fmt(net)}", sty("nr", fontSize=10, fontName=_FONT_BOLD,
                textColor=colors.HexColor(net_color), alignment=TA_RIGHT))]],
            colWidths=[sum(col_w)],
        )
        story.append(net_row)

        # ── Load details (listed underneath the truck table) ──────────────────
        if td.loads:
            story.append(Spacer(1, 2.5*mm))
            story.append(Paragraph("Load Details", s_subhdr))
            ld_rows = [[
                P("Date", s_grp), P("Mine", s_grp), P("Slip #", s_grp),
                P("Tonnes", s_grp_r), P("Rate", s_grp_r), P("Amount", s_grp_r),
            ]]
            tot_tonnes = D0
            tot_amt = D0
            for l in td.loads:
                amt = l.subcontractor_amount_incl_vat if is_vat else l.subcontractor_amount_excl_vat
                if l.tonnes is not None:
                    tot_tonnes += Decimal(str(l.tonnes))
                if amt is not None:
                    tot_amt += Decimal(str(amt))
                ld_rows.append([
                    P(_fmtd(l.load_date), s_dl),
                    P(l.mine_name or "—", s_dl),
                    P(l.slip_number or "—", s_dl),
                    P(_tonnes(l.tonnes) if l.tonnes is not None else "—", s_dl_r),
                    P(_fmt(l.subcontractor_rate) if l.subcontractor_rate is not None else "—", s_dl_r),
                    P(_fmt(amt) if amt is not None else "—", s_dl_r),
                ])
            ld_rows.append([
                P("Total", s_bold), P("", s_dl), P("", s_dl),
                P(_tonnes(tot_tonnes), s_bold_r), P("", s_dl_r), P(_fmt(tot_amt), s_bold_r),
            ])
            ld_tot = len(ld_rows) - 1
            ld_tbl = Table(ld_rows, colWidths=[24*mm, 48*mm, 26*mm, 22*mm, 26*mm, 26*mm])
            ld_tbl.setStyle(TableStyle([
                ("LINEBELOW",    (0, 0), (-1, 0), 0.4, border_c),
                ("LINEABOVE",    (0, ld_tot), (-1, ld_tot), 0.6, border_c),
                ("LEFTPADDING",  (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING",   (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
                ("ROWBACKGROUNDS", (0, 1), (-1, ld_tot - 1), [white, gray_lt]),
                ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ]))
            story.append(ld_tbl)

        # ── Diesel summary (per supplier, listed underneath Load Details) ──────
        if td.diesel_groups:
            story.append(Spacer(1, 3*mm))
            story.append(Paragraph("Diesel Summary", s_subhdr))
            for dg in td.diesel_groups:
                story.append(Spacer(1, 1.5*mm))
                story.append(Paragraph(_diesel_label(dg.supplier_name), s_dsup))
                ds_rows = [[
                    P("Date", s_grp), P("Slip #", s_grp), P("Trans ID", s_grp), P("Invoice #", s_grp),
                    P("Litres", s_grp_r), P("R/Lt", s_grp_r),
                    P("Amt Excl", s_grp_r), P("Admin Fee Incl", s_grp_r), P("Grand Total", s_grp_r),
                ]]
                tot_litres = D0
                for r in dg.rows:
                    if r.litres is not None:
                        tot_litres += Decimal(str(r.litres))
                    ds_rows.append([
                        P(_fmtd(r.fillup_date), s_dl),
                        P(r.depot_slip_number or "—", s_dl),
                        P(r.slip_number or "—", s_dl),
                        P(r.invoice_number or "—", s_dl),
                        P(_tonnes(r.litres) if r.litres is not None else "—", s_dl_r),
                        P(_fmt(r.rate_per_litre) if r.rate_per_litre is not None else "—", s_dl_r),
                        P(_fmt(r.amount_excl), s_dl_r),
                        P(_fmt(r.admin_fee_incl), s_dl_r),
                        P(_fmt(r.grand_total), s_dl_r),
                    ])
                ds_rows.append([
                    P("Total", s_bold), P("", s_dl), P("", s_dl), P("", s_dl),
                    P(_tonnes(tot_litres), s_bold_r), P("", s_dl_r),
                    P(_fmt(dg.tot_excl_admin_fee), s_bold_r),
                    P(_fmt(dg.tot_admin_fee_incl), s_bold_r),
                    P(_fmt(dg.tot_grand_total), s_bold_r),
                ])
                ds_tot = len(ds_rows) - 1
                ds_tbl = Table(ds_rows, colWidths=[20*mm, 18*mm, 18*mm, 16*mm, 14*mm, 16*mm, 24*mm, 26*mm, 26*mm])
                ds_tbl.setStyle(TableStyle([
                    ("LINEBELOW",    (0, 0), (-1, 0), 0.4, border_c),
                    ("LINEABOVE",    (0, ds_tot), (-1, ds_tot), 0.6, border_c),
                    ("LEFTPADDING",  (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING",   (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
                    ("ROWBACKGROUNDS", (0, 1), (-1, ds_tot - 1), [white, gray_lt]),
                    ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
                ]))
                story.append(ds_tbl)

        story.append(Spacer(1, 5*mm))

    # ── Summary ───────────────────────────────────────────────────────────────
    sm = costing.summary
    smry_rows = [
        [P("Summary", s_smry_l), P("Income Excl", s_smry), P("Income Incl", s_smry),
         P("Exp Incl", s_smry), P("Exp Excl", s_smry), P("Net Payable", s_smry)],
        [P("All Trucks", sty("at", fontSize=9)),
         P(_fmt(sm.income_excl_vat), s_bold_r),
         P(_fmt(sm.income_incl_vat), s_bold_r),
         P(_fmt(sm.total_expenses_incl_vat), s_bold_r),
         P(_fmt(sm.total_expenses_excl_vat), s_bold_r),
         P(_fmt(sm.net_payable), s_bold_r)],
    ]
    smry_col_w = [42*mm, 26*mm, 26*mm, 26*mm, 26*mm, 26*mm]
    smry_tbl = Table(smry_rows, colWidths=smry_col_w)
    smry_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR",    (0, 0), (-1, 0), white),
        ("FONTNAME",     (0, 0), (-1, 0), _FONT_BOLD),
        ("ALIGN",        (1, 0), (-1, -1), "RIGHT"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("GRID",         (0, 0), (-1, -1), 0.4, border_c),
        ("BACKGROUND",   (0, 1), (-1, 1), gray_lt),
        ("FONTNAME",     (0, 1), (-1, 1), _FONT_BOLD),
    ]))
    story.append(smry_tbl)

    doc.build(story)
    return buf.getvalue()


# ── Excel ─────────────────────────────────────────────────────────────────────

def generate_costing_excel(costing, entity) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from app.services.pdf_generator import _load_logo

    wb = Workbook()
    ws = wb.active
    ws.title = f"{MONTH_NAMES[costing.month]} {costing.year}"
    # Per-load detail rows are grouped under each truck's "Loads" summary row.
    # summaryBelow=False puts the expand/collapse control on the summary row above.
    ws.sheet_properties.outlinePr.summaryBelow = False

    # Style helpers
    def fill(hex_str):
        return PatternFill("solid", fgColor=hex_str.lstrip("#"))

    def font(bold=False, color="111111", size=10):
        return Font(bold=bold, color=color, size=size)

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin = Side(style="thin", color="D1D5DB")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    COL_W = [38, 16, 16, 16, 16]
    for i, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # F/G/H/I are only used by the per-supplier Diesel Summary tables below.
    for letter, w in zip(("F", "G", "H", "I"), (14, 16, 16, 16)):
        ws.column_dimensions[letter].width = w

    row = 1

    # ── Letterhead image ──────────────────────────────────────────────────────
    from openpyxl.drawing.image import Image as XLImage
    lh_url  = getattr(entity, "letterhead_url",  None)
    lh_path = getattr(entity, "letterhead_path", None)
    lh_added = False
    if lh_url or lh_path:
        lh_bytes = _load_logo(lh_url, lh_path)
        if lh_bytes:
            try:
                from PIL import Image as PILImage
                pil_img = PILImage.open(io.BytesIO(lh_bytes))
                orig_w, orig_h = pil_img.size
                target_w_px = 600
                scale = target_w_px / orig_w
                target_h_px = int(orig_h * scale)
                # Convert to PNG for openpyxl
                out = io.BytesIO()
                pil_img.save(out, format="PNG")
                out.seek(0)
                xl_img = XLImage(out)
                xl_img.width  = target_w_px
                xl_img.height = target_h_px
                ws.add_image(xl_img, "A1")
                # Reserve rows for the image (approx 15px per row)
                img_rows = max(1, target_h_px // 15)
                for _ in range(img_rows):
                    ws.append([])
                row += img_rows
                lh_added = True
            except Exception:
                pass

    if not lh_added:
        ent_name = getattr(entity, "name", "") or ""
        ws.cell(row=row, column=1, value=ent_name).font = font(bold=True, size=14)
        ws.row_dimensions[row].height = 22
        row += 1

    # ── Title ─────────────────────────────────────────────────────────────────
    period   = f"{MONTH_NAMES[costing.month]} {costing.year}"
    sub_name = costing.subcontractor.name if costing.subcontractor else ""

    c = ws.cell(row=row, column=1, value=f"Subcontractor Costing — {sub_name}")
    c.font = font(bold=True, size=12)
    row += 1
    c2 = ws.cell(row=row, column=1, value=f"Period: {period}")
    c2.font = font(color="6B7280", size=10)
    row += 2

    HDR_FILL   = fill("4B5563")
    DARK_FILL  = fill("1F2937")
    ALT_FILL   = fill("F9FAFB")
    GRP_FILL   = fill("F3F4F6")
    TOTAL_FILL = fill("374151")
    HDR_FONT   = font(bold=True, color="FFFFFF", size=9)
    LBL_FONT   = font(size=9)
    VAL_FONT   = font(size=9)
    TOT_FONT   = font(bold=True, size=9)
    TOT_WHITE  = font(bold=True, color="FFFFFF", size=9)
    GRP_FONT   = font(bold=True, color="6B7280", size=8)

    def write_row(values, fnt=None, fills=None, aligns=None, row_h=None):
        nonlocal row
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            if fnt:
                c.font = fnt
            if fills and fills[col - 1]:
                c.fill = fills[col - 1]
            if aligns and aligns[col - 1]:
                c.alignment = aligns[col - 1]
            else:
                c.alignment = align(h="right" if col > 1 else "left")
            c.border = border_all
        if row_h:
            ws.row_dimensions[row].height = row_h
        row += 1

    headers = ["", "Income Excl VAT", "Income Incl VAT", "Expenses Incl VAT", "Expenses Excl VAT"]
    is_vat = costing.is_vat_registered

    def fmt_val(v):
        try:
            return round(float(v), 2)
        except (TypeError, ValueError):
            return None

    for td in costing.trucks:
        t = td.truck
        fleet = f" · {t.fleet_number}" if t.fleet_number else ""
        reg_label = f"Truck: {t.registration or ''}{fleet}"

        # Truck header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=reg_label)
        c.font = font(bold=True, size=11)
        c.fill = fill("E5E7EB")
        c.alignment = align()
        ws.row_dimensions[row].height = 18
        row += 1

        # Column headers
        write_row(headers, fnt=HDR_FONT,
                  fills=[HDR_FILL] * 5,
                  aligns=[align(h="left")] + [align(h="right")] * 4,
                  row_h=16)

        inc_e = fmt_val(td.income_excl_vat)
        inc_i = fmt_val(td.income_incl_vat)
        loads_inc_e = fmt_val(td.loads_income_excl_vat)
        loads_inc_i = fmt_val(td.loads_income_incl_vat)
        net   = fmt_val(td.net_payable)

        # Column letters: income excl VAT always lands in B, incl VAT in C
        # (only populated when the entity is VAT-registered); expenses always
        # in D (incl) and E (excl).
        EXCL_L  = get_column_letter(2)
        INCL_L  = get_column_letter(3)
        EXP_I_L = get_column_letter(4)
        EXP_E_L = get_column_letter(5)
        exp_incl_cells = []  # cells summed into the Expenses Incl VAT total
        exp_excl_cells = []  # cells summed into the Expenses Excl VAT total

        # ── Loads summary row — income is a SUM of the per-load detail rows
        #    grouped (collapsed) directly beneath it ──────────────────────────
        loads_row = row
        n_loads   = len(td.loads)
        if n_loads:
            excl_val = f"=SUM({EXCL_L}{loads_row + 1}:{EXCL_L}{loads_row + n_loads})"
            incl_val = f"=SUM({INCL_L}{loads_row + 1}:{INCL_L}{loads_row + n_loads})" if is_vat else None
        else:
            excl_val = loads_inc_e
            incl_val = loads_inc_i if is_vat else None
        write_row(
            [f"Loads ({n_loads})", excl_val, incl_val, None, None],
            fnt=LBL_FONT, fills=[None]*5, row_h=15,
        )
        inc_excl_cells = [f"{EXCL_L}{loads_row}"]  # cells summed into the Income Excl VAT total
        inc_incl_cells = [f"{INCL_L}{loads_row}"]  # cells summed into the Income Incl VAT total

        # Per-load detail rows (collapsed by default; expand via the +/- control)
        detail_font = font(color="6B7280", size=8)
        for l in td.loads:
            amt_excl = fmt_val(l.subcontractor_amount_excl_vat)
            amt_incl = fmt_val(l.subcontractor_amount_incl_vat)
            write_row(
                [_load_label(l),
                 amt_excl,
                 amt_incl if is_vat else None,
                 None, None],
                fnt=detail_font, fills=[None]*5, row_h=14,
            )
            ws.row_dimensions[row - 1].outline_level = 1
            ws.row_dimensions[row - 1].hidden = True

        # Manual income lines — added in the costing module only
        for mi in (td.manual_incomes or []):
            write_row(
                [mi.description, fmt_val(mi.amount_excl_vat), fmt_val(mi.amount_incl_vat) if is_vat else None, None, None],
                fnt=LBL_FONT, fills=[None]*5, row_h=15,
            )
            inc_excl_cells.append(f"{EXCL_L}{row - 1}")
            inc_incl_cells.append(f"{INCL_L}{row - 1}")

        # Admin fee
        write_row(["Admin Fee", None, None, fmt_val(td.admin_fee), None], fnt=LBL_FONT, row_h=15)
        exp_incl_cells.append(f"{EXP_I_L}{row - 1}")

        # Diesel groups
        for dg in (td.diesel_groups or []):
            write_row([_diesel_label(dg.supplier_name), None, None, None, fmt_val(dg.tot_excl_admin_fee)], fnt=LBL_FONT, row_h=15)
            exp_excl_cells.append(f"{EXP_E_L}{row - 1}")
            write_row([f"{_diesel_label(dg.supplier_name)} Admin Fee", None, None, fmt_val(dg.tot_admin_fee_incl), None], fnt=LBL_FONT, row_h=15)
            exp_incl_cells.append(f"{EXP_I_L}{row - 1}")

        # Supplier invoices
        if td.supplier_invoices:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            c = ws.cell(row=row, column=1, value="Supplier Invoices")
            c.font = GRP_FONT
            c.fill = GRP_FILL
            c.alignment = align()
            row += 1
            for inv in td.supplier_invoices:
                sup = inv.supplier_name or f"Supplier #{inv.supplier_id}"
                if inv.vat_applicable:
                    write_row([sup, None, None, fmt_val(inv.amount), None], fnt=LBL_FONT, row_h=15)
                    exp_incl_cells.append(f"{EXP_I_L}{row - 1}")
                else:
                    write_row([sup, None, None, None, fmt_val(inv.amount)], fnt=LBL_FONT, row_h=15)
                    exp_excl_cells.append(f"{EXP_E_L}{row - 1}")

        # Totals — built from Excel SUM formulas so the component cells are
        # visible: income sums the Loads row plus any manual income lines,
        # expenses sum their line items.
        def _sum(cells):
            return f"=SUM({','.join(cells)})" if cells else 0
        inc_total_excl = _sum(inc_excl_cells)
        inc_total_incl = _sum(inc_incl_cells) if is_vat else None
        write_row(
            ["Totals",
             inc_total_excl,
             inc_total_incl,
             _sum(exp_incl_cells), _sum(exp_excl_cells)],
            fnt=TOT_WHITE,
            fills=[TOTAL_FILL] * 5,
            row_h=16,
        )

        # Net payable
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=f"Net Payable: {_fmt(net)}")
        net_color = "15803D" if (net or 0) >= 0 else "DC2626"
        c.font = font(bold=True, color=net_color, size=10)
        c.alignment = align(h="right")
        ws.row_dimensions[row].height = 16
        row += 1

        # ── Diesel summary (per supplier, listed underneath the truck block) ───
        if td.diesel_groups:
            row += 1
            c = ws.cell(row=row, column=1, value="Diesel Summary")
            c.font = font(bold=True, size=10)
            row += 1

            ds_headers = ["Date", "Slip #", "Trans ID", "Invoice #", "Litres", "R/Lt",
                          "Amt Excl", "Admin Fee Incl", "Grand Total"]
            ds_align = lambda col: align(h="left" if col <= 4 else "right")

            for dg in td.diesel_groups:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
                c = ws.cell(row=row, column=1, value=_diesel_label(dg.supplier_name))
                c.font = font(bold=True, size=9.5)
                c.fill = fill("F3F4F6")
                row += 1

                for col, val in enumerate(ds_headers, 1):
                    cc = ws.cell(row=row, column=col, value=val)
                    cc.font = font(bold=True, color="6B7280", size=8)
                    cc.alignment = ds_align(col)
                    cc.border = Border(bottom=thin)
                row += 1

                tot_litres = 0.0
                for r in dg.rows:
                    lit = fmt_val(r.litres)
                    if lit is not None:
                        tot_litres += lit
                    vals = [_fmtd(r.fillup_date), r.depot_slip_number or "—", r.slip_number or "—", r.invoice_number or "—",
                            lit, fmt_val(r.rate_per_litre), fmt_val(r.amount_excl),
                            fmt_val(r.admin_fee_incl), fmt_val(r.grand_total)]
                    for col, val in enumerate(vals, 1):
                        cc = ws.cell(row=row, column=col, value=val)
                        cc.font = font(color="6B7280", size=8)
                        cc.alignment = ds_align(col)
                    row += 1

                tot_vals = ["Total", None, None, None, round(tot_litres, 2), None,
                            fmt_val(dg.tot_excl_admin_fee), fmt_val(dg.tot_admin_fee_incl), fmt_val(dg.tot_grand_total)]
                for col, val in enumerate(tot_vals, 1):
                    cc = ws.cell(row=row, column=col, value=val)
                    cc.font = font(bold=True, size=8)
                    cc.alignment = ds_align(col)
                    cc.border = Border(top=thin)
                row += 2

        row += 1

    # ── Summary ───────────────────────────────────────────────────────────────
    sm = costing.summary
    sum_headers = ["Summary", "Income Excl VAT", "Income Incl VAT", "Exp Incl VAT", "Exp Excl VAT"]
    write_row(sum_headers, fnt=HDR_FONT,
              fills=[DARK_FILL] * 5,
              aligns=[align(h="left")] + [align(h="right")] * 4,
              row_h=16)
    write_row(
        ["All Trucks",
         fmt_val(sm.income_excl_vat),
         fmt_val(sm.income_incl_vat),
         fmt_val(sm.total_expenses_incl_vat),
         fmt_val(sm.total_expenses_excl_vat)],
        fnt=TOT_FONT,
        fills=[ALT_FILL] * 5,
        row_h=15,
    )

    # Net payable summary
    net_sm = fmt_val(sm.net_payable)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=f"Total Net Payable: {_fmt(net_sm)}")
    c.font = font(bold=True, color="15803D" if (net_sm or 0) >= 0 else "DC2626", size=11)
    c.alignment = align(h="right")
    ws.row_dimensions[row].height = 18
    row += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
