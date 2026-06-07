"""
Statement generators — safetec_core

PDF (ReportLab) and Excel (openpyxl) statement documents that share the
invoice look & feel: entity letterhead as the full-width header, entity brand
colour for headings/totals, Arial font, neutral gray column headers.

Layout (generic statement):
    DESCRIPTION | DATE PAID | PAID | OUTSTANDING
  - invoice (outstanding) lines carry a positive amount
  - payment lines carry a negative amount
  - totals: PAID (sum of payments) and AMOUNT DUE
"""

import io
from decimal import Decimal

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    Image as RLImage,
)

from app.services.pdf_generator import (
    _load_logo, _hex_to_color, format_currency, format_date,
    _FONT_NORMAL, _FONT_BOLD, _FONT_ITALIC,
)


# ── Shared helpers ────────────────────────────────────────────────────────────

def _sorted_lines(stmt):
    return sorted(stmt.lines, key=lambda l: getattr(l, "sort_order", 0) or 0)


def _totals(lines):
    outstanding = Decimal("0")
    paid = Decimal("0")
    for l in lines:
        amt = Decimal(str(l.amount or 0))
        if amt < 0:
            paid += amt
        else:
            outstanding += amt
    return outstanding, paid, outstanding + paid


def _invoice_desc(line):
    if line.description:
        return line.description
    parts = []
    if line.line_date:
        parts.append(format_date(line.line_date))
    if line.invoice_number:
        parts.append(line.invoice_number)
    return " - ".join(parts)


# ── PDF ───────────────────────────────────────────────────────────────────────

def generate_statement_pdf(stmt, entity, customer) -> bytes:
    accent     = _hex_to_color(getattr(entity, "primary_color", None) or "#1a1a2e")
    black      = colors.HexColor("#111111")
    gray_dark  = colors.HexColor("#374151")
    gray_mid   = colors.HexColor("#6b7280")
    gray_light = colors.HexColor("#f3f4f6")
    white      = colors.white
    divider    = colors.HexColor("#e5e7eb")
    col_hdr_bg = colors.HexColor("#4B5563")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=10*mm, leftMargin=10*mm, topMargin=10*mm, bottomMargin=16*mm,
    )
    CONTENT_W = 190  # mm (A4 width 210 − 2×10 margins)
    story = []

    def st(name, **kw):
        defaults = dict(fontName=_FONT_NORMAL, fontSize=9, textColor=black, leading=12)
        defaults.update(kw)
        return ParagraphStyle(name, **defaults)

    s_company_name  = st("co_name", fontSize=10, fontName=_FONT_BOLD, leading=14)
    s_company_sub   = st("co_sub",  fontSize=7.5, textColor=gray_mid, leading=10)
    s_contact_val   = st("ct_val",  fontSize=8,   textColor=gray_dark, leading=11)
    s_section_label = st("sec_lbl", fontSize=8,   fontName=_FONT_BOLD, textColor=gray_mid)
    s_client_name   = st("cl_name", fontSize=11,  fontName=_FONT_BOLD, leading=15)
    s_client_detail = st("cl_det",  fontSize=9,   textColor=gray_dark, leading=12)
    s_inv_label     = st("inv_lbl", fontSize=8,   textColor=gray_mid, alignment=TA_RIGHT, leading=11)
    s_inv_value     = st("inv_val", fontSize=9,   fontName=_FONT_BOLD, alignment=TA_RIGHT, leading=12)
    s_col_header    = st("col_hdr", fontSize=9,   fontName=_FONT_BOLD, textColor=accent, leading=12)
    s_col_hdr_c     = st("col_hdr_c", fontSize=9, fontName=_FONT_BOLD, textColor=accent, alignment=TA_CENTER, leading=12)
    s_col_hdr_r     = st("col_hdr_r", fontSize=9, fontName=_FONT_BOLD, textColor=accent, alignment=TA_RIGHT, leading=12)
    s_line_desc     = st("ln_desc", fontSize=9,   textColor=gray_dark, leading=12)
    s_line_ctr      = st("ln_ctr",  fontSize=9,   textColor=gray_dark, alignment=TA_CENTER, leading=12)
    s_line_num      = st("ln_num",  fontSize=9,   textColor=gray_dark, alignment=TA_RIGHT, leading=12)
    s_pay_num       = st("pay_num", fontSize=9,   textColor=colors.HexColor("#15803d"), alignment=TA_RIGHT, leading=12)
    s_total_label   = st("tot_lbl", fontSize=9,   textColor=gray_dark, alignment=TA_RIGHT, leading=12)
    s_total_value   = st("tot_val", fontSize=9,   fontName=_FONT_BOLD, alignment=TA_RIGHT, leading=12)
    s_grand_label   = st("gr_lbl",  fontSize=11,  fontName=_FONT_BOLD, textColor=white, alignment=TA_RIGHT, leading=14)
    s_grand_value   = st("gr_val",  fontSize=11,  fontName=_FONT_BOLD, textColor=white, alignment=TA_RIGHT, leading=14)
    s_bank_title    = st("bk_title", fontSize=8,  fontName=_FONT_BOLD, textColor=gray_dark, leading=11)
    s_bank_detail   = st("bk_det",  fontSize=8,   textColor=gray_mid, leading=11)
    s_footer        = st("footer",  fontSize=7,   textColor=gray_mid, alignment=TA_CENTER, leading=9)

    # ── Header: letterhead (preferred) or logo + company info ─────────────────
    letterhead_img = None
    lh_bytes = _load_logo(getattr(entity, "letterhead_url", None), getattr(entity, "letterhead_path", None))
    if lh_bytes:
        try:
            letterhead_img = RLImage(io.BytesIO(lh_bytes), width=CONTENT_W*mm, height=130*mm, kind="proportional")
            letterhead_img.hAlign = "CENTER"
        except Exception:
            letterhead_img = None

    if letterhead_img:
        story.append(letterhead_img)
        story.append(Spacer(1, 10*mm))
    else:
        logo_img = None
        logo_bytes = _load_logo(getattr(entity, "logo_url", None), getattr(entity, "logo_path", None))
        if logo_bytes:
            try:
                logo_img = RLImage(io.BytesIO(logo_bytes), width=110*mm, height=45*mm, kind="proportional")
            except Exception:
                logo_img = None

        left_col = [Paragraph(entity.trading_name or entity.name, s_company_name)]
        if entity.registration_number:
            left_col.append(Paragraph(f"Reg {entity.registration_number}", s_company_sub))
        if entity.vat_number:
            left_col.append(Paragraph(f"VAT No: {entity.vat_number}", s_company_sub))
        if entity.phone:
            left_col.append(Paragraph(f"Tel:  {entity.phone}", s_contact_val))
        if entity.email:
            left_col.append(Paragraph(entity.email, s_contact_val))
        if entity.address:
            for line in [l.strip() for l in entity.address.replace(",", "\n").split("\n") if l.strip()]:
                left_col.append(Paragraph(line, s_contact_val))

        right_col = []
        if logo_img:
            logo_cell = Table([[logo_img]], colWidths=[100*mm])
            logo_cell.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ]))
            right_col = [logo_cell]

        header_table = Table([[left_col, right_col]], colWidths=[80*mm, 110*mm], hAlign="LEFT")
        header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(header_table)
        story.append(HRFlowable(width="100%", thickness=0.5, color=divider))
        story.append(Spacer(1, 4*mm))

    lines = _sorted_lines(stmt)
    outstanding, paid, amount_due = _totals(lines)

    # ── Statement-to block + meta (date / amount due) ─────────────────────────
    bill_lines = [Paragraph("STATEMENT TO", s_section_label),
                  Paragraph(customer.name if customer else "—", s_client_name)]
    if customer:
        if customer.registration_number:
            bill_lines.append(Paragraph(f"Reg {customer.registration_number}", s_client_detail))
        if customer.address:
            for addr_line in customer.address.split("\n"):
                if addr_line.strip():
                    bill_lines.append(Paragraph(addr_line.strip(), s_client_detail))
        if customer.city:
            postal = f"{customer.city}, {customer.postal_code}" if customer.postal_code else customer.city
            bill_lines.append(Paragraph(postal, s_client_detail))
        if customer.vat_number:
            bill_lines.append(Paragraph(f"VAT NO:  {customer.vat_number}", s_client_detail))

    meta_rows = [
        [Paragraph("Statement Date", s_inv_label), Paragraph(format_date(stmt.statement_date), s_inv_value)],
        [Paragraph("Amount Due",     s_inv_label), Paragraph(format_currency(amount_due), s_inv_value)],
    ]
    meta_table = Table(meta_rows, colWidths=[30*mm, 36*mm])
    meta_table.setStyle(TableStyle([
        ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))

    bill_section = Table([[bill_lines, meta_table]], colWidths=[118*mm, 72*mm])
    bill_section.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(bill_section)
    story.append(Spacer(1, 5*mm))

    # ── Lines table ───────────────────────────────────────────────────────────
    col_widths = [98*mm, 26*mm, 30*mm, 36*mm]
    table_rows = [[
        Paragraph("DESCRIPTION", s_col_header),
        Paragraph("DATE PAID",   s_col_hdr_c),
        Paragraph("PAID",        s_col_hdr_r),
        Paragraph("OUTSTANDING", s_col_hdr_r),
    ]]
    item_row_idxs = []
    for line in lines:
        amt = Decimal(str(line.amount or 0))
        idx = len(table_rows)
        if amt < 0:
            table_rows.append([
                Paragraph(line.description or "", s_line_desc),
                Paragraph(format_date(line.line_date) if line.line_date else "", s_line_ctr),
                Paragraph(format_currency(amt), s_pay_num),
                Paragraph("", s_line_num),
            ])
        else:
            table_rows.append([
                Paragraph(_invoice_desc(line), s_line_desc),
                Paragraph("", s_line_ctr),
                Paragraph("", s_line_num),
                Paragraph(format_currency(amt), s_line_num),
            ])
        item_row_idxs.append(idx)

    while len(table_rows) < 9:
        table_rows.append(["", "", "", ""])

    items_table = Table(table_rows, colWidths=col_widths)
    row_styles = [
        ("BACKGROUND",    (0, 0), (-1, 0), col_hdr_bg),
        ("TOPPADDING",    (0, 0), (-1, 0), 7), ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
        ("TOPPADDING",    (0, 1), (-1, -1), 5), ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6), ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW",     (0, 0), (-1, -1), 0.3, divider),
    ]
    for r in item_row_idxs:
        if r % 2 == 0:
            row_styles.append(("BACKGROUND", (0, r), (-1, r), gray_light))
    items_table.setStyle(TableStyle(row_styles))
    story.append(items_table)
    story.append(Spacer(1, 4*mm))

    # ── Banking + totals ──────────────────────────────────────────────────────
    bank_content = []
    if entity.bank_account_number:
        bank_content.append(Paragraph("Banking Details", s_bank_title))
        bank_content.append(Spacer(1, 2*mm))
        if entity.bank_name:
            bank_content.append(Paragraph(f"Bank:  {entity.bank_name}", s_bank_detail))
        if entity.bank_branch:
            bank_content.append(Paragraph(f"Branch:  {entity.bank_branch}", s_bank_detail))
        if entity.bank_branch_code:
            bank_content.append(Paragraph(f"Branch Code:  {entity.bank_branch_code}", s_bank_detail))
        bank_content.append(Paragraph(f"Account No:  {entity.bank_account_number}", s_bank_detail))
        if entity.bank_reference:
            bank_content.append(Paragraph(f"Reference:  {entity.bank_reference}", s_bank_detail))

    totals_data = [
        [Paragraph("Paid",       s_total_label), Paragraph(format_currency(paid), s_total_value)],
        [Paragraph("Amount Due", s_grand_label), Paragraph(format_currency(amount_due), s_grand_value)],
    ]
    totals_right = Table(totals_data, colWidths=[30*mm, 36*mm])
    totals_right.setStyle(TableStyle([
        ("BACKGROUND",    (0, 1), (-1, 1), col_hdr_bg),
        ("TOPPADDING",    (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4), ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ("LINEABOVE",     (0, 0), (-1, 0), 0.5, divider),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))

    bottom_table = Table([[bank_content, totals_right]], colWidths=[118*mm, 72*mm])
    bottom_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(bottom_table)
    story.append(Spacer(1, 6*mm))

    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=divider))
    story.append(Spacer(1, 2*mm))
    footer_text = entity.name
    if entity.vat_number:
        footer_text += f"  |  VAT: {entity.vat_number}"
    story.append(Paragraph(footer_text, s_footer))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


# ── Excel ─────────────────────────────────────────────────────────────────────

def generate_statement_excel(stmt, entity, customer) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    brand_hex = (getattr(entity, "primary_color", None) or "#1A1A2E").lstrip("#").upper()
    if len(brand_hex) == 3:
        brand_hex = "".join(c * 2 for c in brand_hex)
    brand_fill = PatternFill(start_color=brand_hex, end_color=brand_hex, fill_type="solid")
    grand_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
    white_bold = Font(bold=True, color="FFFFFF")
    thin       = Side(style="thin", color="D0D5DD")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt  = '#,##0.00'

    lines = _sorted_lines(stmt)
    outstanding, paid, amount_due = _totals(lines)

    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    row = 1

    # Letterhead image (floats over the cells) — reserve rows beneath it
    lh_bytes = _load_logo(getattr(entity, "letterhead_url", None), getattr(entity, "letterhead_path", None))
    if not lh_bytes:
        lh_bytes = _load_logo(getattr(entity, "logo_url", None), getattr(entity, "logo_path", None))
    if lh_bytes:
        try:
            from PIL import Image as PILImage
            pil = PILImage.open(io.BytesIO(lh_bytes))
            ratio = pil.height / pil.width if pil.width else 0.3
            disp_w = 820
            disp_h = int(disp_w * ratio)
            img = XLImage(io.BytesIO(lh_bytes))
            img.width, img.height = disp_w, disp_h
            ws.add_image(img, "A1")
            row = max(2, int(disp_h / 18) + 2)
        except Exception:
            pass

    def put(r, c, val, *, font=None, fill=None, align=None, fmt=None, bordered=False):
        cell = ws.cell(row=r, column=c, value=val)
        if font:  cell.font = font
        if fill:  cell.fill = fill
        if align: cell.alignment = align
        if fmt:   cell.number_format = fmt
        if bordered: cell.border = border
        return cell

    # Customer block + meta
    name_row = row
    put(row, 1, customer.name if customer else "", font=Font(bold=True, size=13)); row += 1
    if customer and customer.registration_number:
        put(row, 1, f"REG: {customer.registration_number}"); row += 1
    if customer and customer.address:
        for line in [l.strip() for l in customer.address.split("\n") if l.strip()]:
            put(row, 1, line); row += 1
    if customer and customer.city:
        put(row, 1, customer.city); row += 1
    if customer and customer.postal_code:
        put(row, 1, str(customer.postal_code)); row += 1
    if customer and customer.vat_number:
        put(row, 1, f"VAT NO: {customer.vat_number}"); row += 1

    # Date + amount due to the right of the customer name
    put(name_row, 3, "Statement Date", font=Font(bold=True, color="667085"), align=Alignment(horizontal="right"))
    put(name_row, 4, format_date(stmt.statement_date), align=Alignment(horizontal="right"))
    put(name_row + 1, 3, "Amount Due", font=Font(bold=True, color="667085"), align=Alignment(horizontal="right"))
    put(name_row + 1, 4, float(amount_due), align=Alignment(horizontal="right"), fmt=money_fmt)

    row += 1
    # Header row (brand colour)
    headers = ["DESCRIPTION", "DATE PAID", "PAID", "OUTSTANDING"]
    for c, h in enumerate(headers, start=1):
        put(row, c, h, font=white_bold, fill=brand_fill, bordered=True,
            align=Alignment(horizontal="left" if c == 1 else ("center" if c == 2 else "right")))
    row += 1

    for line in lines:
        amt = float(line.amount or 0)
        if amt < 0:
            put(row, 1, line.description or "", bordered=True)
            put(row, 2, format_date(line.line_date) if line.line_date else "", align=Alignment(horizontal="center"), bordered=True)
            put(row, 3, amt, fmt=money_fmt, align=Alignment(horizontal="right"), bordered=True)
            put(row, 4, None, bordered=True)
        else:
            put(row, 1, _invoice_desc(line), bordered=True)
            put(row, 2, None, bordered=True)
            put(row, 3, None, bordered=True)
            put(row, 4, amt, fmt=money_fmt, align=Alignment(horizontal="right"), bordered=True)
        row += 1

    # Totals
    put(row, 2, "PAID", font=Font(bold=True), align=Alignment(horizontal="right"))
    put(row, 3, float(paid), fmt=money_fmt, font=Font(bold=True), align=Alignment(horizontal="right"))
    row += 1
    put(row, 2, "AMOUNT DUE", font=white_bold, fill=grand_fill, align=Alignment(horizontal="right"))
    put(row, 3, float(amount_due), fmt=money_fmt, font=white_bold, fill=grand_fill, align=Alignment(horizontal="right"))
    row += 2

    # Banking details
    if entity.bank_account_number:
        put(row, 1, "Banking Details", font=Font(bold=True)); row += 1
        if entity.name:                put(row, 1, entity.name); row += 1
        if entity.bank_name:           put(row, 1, entity.bank_name); row += 1
        if entity.bank_branch:         put(row, 1, f"Branch {entity.bank_branch}"); row += 1
        if entity.bank_branch_code:    put(row, 1, f"Branch Code: {entity.bank_branch_code}"); row += 1
        put(row, 1, f"Account No: {entity.bank_account_number}"); row += 1
        if entity.bank_reference:      put(row, 1, f"Reference: {entity.bank_reference}"); row += 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
