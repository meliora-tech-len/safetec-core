"""
Budget export generators — PDF (reportlab) and Excel (openpyxl).
Both include the entity letterhead/logo as a header, mirroring the costing
exports, and lay the document out the way the Budgets page shows it:
summary cards, Bank Info Summary, then every section's grid.

The generators take the detail-loaded Budget (sections → lines → values,
bank_rows, entity) plus the window of (month, year) columns the page shows —
computed by the route via _budget_window, so the export always matches the
grid on screen.
"""

import io
import re
from decimal import Decimal

from app.core.constants import (
    MONTH_NAMES_1 as MONTH_NAMES,
    MONTHS_SHORT_1 as MONTHS_SHORT,
)

D0 = Decimal("0")

# ── Frontend mirrors (keep in step with BudgetsPage.jsx) ──────────────────────

# Safetec's summary splits expenses in two: what the trucks cost, and the
# personal/non-truck spend — shown again as "Personal Expenses" and taken off
# Profit to land on Actual Profit.
SFT_PERSONAL_SECTIONS = {"PERSONAL NOT ON TRUCK COSTING", "BUSINESS NOT ON TRUCK COSTING"}

# Entities whose summary is the two per-month-group cards instead of the
# single window total.
MONTH_SPLIT_ENTITIES = {"OBHI", "BTP", "TP"}

# A bank row labelled "Cash Flow Total" is computed (sum of the account rows),
# unless an amount was typed to pin it.
CASH_FLOW_TOTAL_RE = re.compile(r"cash\s*flow\s*total", re.IGNORECASE)

# "vat back trailer" is derived: a line counts when its amount for the base
# month matches the VAT-back figure noted in its own name.
VAT_BACK_RE = re.compile(r"VAT\s*BACK\s*R\s*([\d.,\s]+)", re.IGNORECASE)
VAT_BACK_TOLERANCE = 5


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _fmt(v) -> str:
    # NBSP after the R (project currency convention) — also stops ReportLab
    # wrapping "R" away from its figure in narrow columns.
    try:
        return f"R {float(v):,.2f}"
    except (TypeError, ValueError):
        return "—"


def _vat_back_from_line(line, amount: float):
    m = VAT_BACK_RE.search(line.name or "")
    if not m:
        return None
    raw = m.group(1).replace(" ", "").replace(",", "").rstrip(".")
    try:
        stated = float(raw)
    except ValueError:
        return None
    if not stated:
        return None
    return {"stated": stated, "amount": amount, "name": line.name} \
        if abs(stated - amount) <= VAT_BACK_TOLERANCE else None


def _value_map(line) -> dict:
    return {(v.month, v.year): v for v in (line.values or [])}


def _sorted_lines(section) -> list:
    """Alphabetical by name — the page's default order."""
    return sorted(section.lines, key=lambda l: ((l.name or "").upper(), l.id))


def compute_budget_stats(budget, months, code: str) -> dict:
    """Every derived figure the page shows, computed the same way it computes
    them: window totals, the base-month profit statement, the OBHI/BTP/TP
    month-split groups and the Bank Info Summary block."""
    code = (code or "").upper()
    is_sft = code == "SFT"
    month_set = set(months)
    base = months[0]

    income = expenses_due = expenses_paid = 0.0
    income_by = {k: 0.0 for k in months}
    expense_by = {k: 0.0 for k in months}
    b_income = b_expenses = b_personal = 0.0
    vat_back_hits = []

    for s in budget.sections:
        is_income = s.section_type == "income"
        is_personal = (s.name or "").upper() in SFT_PERSONAL_SECTIONS
        for l in s.lines:
            vmap = _value_map(l)
            for (m, y), v in vmap.items():
                if (m, y) not in month_set:
                    continue
                due, paid = _num(v.amount_due), _num(v.amount_paid)
                if is_income:
                    income += due
                    income_by[(m, y)] += due
                    if (m, y) == base:
                        b_income += due
                else:
                    expenses_due += due
                    expenses_paid += paid
                    expense_by[(m, y)] += due + paid
                    if (m, y) == base:
                        b_expenses += due + paid
                        if is_personal:
                            b_personal += due + paid
            if not is_income:
                bv = vmap.get(base)
                if bv is not None:
                    hit = _vat_back_from_line(l, _num(bv.amount_due) + _num(bv.amount_paid))
                    if hit:
                        vat_back_hits.append(hit)

    b_profit = b_income - b_expenses
    stats = {
        "is_sft": is_sft,
        "base_label": f"{MONTH_NAMES[base[0]]} {base[1]}",
        "totals": {
            "income": income, "expenses_due": expenses_due, "expenses_paid": expenses_paid,
            "left_over": income - expenses_due - expenses_paid,
        },
        "base": {
            "income": b_income, "expenses": b_expenses, "profit": b_profit,
            "personal": b_personal, "actual_profit": b_profit - b_personal,
        },
        "month_split": None,
        "bank": None,
    }

    if code in MONTH_SPLIT_ENTITIES and len(months) >= 3:
        m0, m1, m2 = months[0], months[1], months[2]
        g1_income, g1_exp = income_by[m0], expense_by[m0]
        g2_income = income_by[m1]
        g2_exp = expense_by[m1] + expense_by[m2]
        stats["month_split"] = [
            {"title": f"{MONTHS_SHORT[m0[0]]} {m0[1]}",
             "sub": f"{MONTHS_SHORT[m0[0]]} income − {MONTHS_SHORT[m0[0]]} expenses",
             "income": g1_income, "expenses": g1_exp, "left_over": g1_income - g1_exp},
            {"title": f"{MONTHS_SHORT[m1[0]]} {m1[1]}",
             "sub": f"{MONTHS_SHORT[m1[0]]} income − {MONTHS_SHORT[m1[0]]} & {MONTHS_SHORT[m2[0]]} expenses",
             "income": g2_income, "expenses": g2_exp, "left_over": g2_income - g2_exp},
        ]

    rows = budget.bank_rows or []
    if rows:
        bank_rows = [r for r in rows if r.kind == "bank"]
        cash_flow_row = next((r for r in bank_rows if CASH_FLOW_TOTAL_RE.search(r.label or "")), None)
        accounts = [r for r in bank_rows if r is not cash_flow_row]
        cash_flow_calc = sum(_num(r.amount) for r in accounts)
        to_be_paid = [r for r in rows if r.kind == "to_be_paid"]

        profit = _num(budget.bank_profit_override) if budget.bank_profit_override is not None else b_profit
        vat_back_pulled = sum(h["amount"] for h in vat_back_hits)
        vat_back = _num(budget.vat_back_trailer) if budget.vat_back_trailer is not None else vat_back_pulled
        profit_excl = (_num(budget.profit_excl_vat_back_override)
                       if budget.profit_excl_vat_back_override is not None
                       else profit + vat_back)
        stats["bank"] = {
            "accounts": accounts,
            "cash_flow_row": cash_flow_row,
            "cash_flow_total": (_num(cash_flow_row.amount)
                                if cash_flow_row is not None and cash_flow_row.amount is not None
                                else cash_flow_calc),
            "to_be_paid": to_be_paid,
            "to_be_paid_total": sum(_num(r.amount) for r in to_be_paid),
            "profit": profit,
            "vat_back": vat_back,
            "profit_excl_vat_back": profit_excl,
        }
    return stats


# ── PDF ───────────────────────────────────────────────────────────────────────

def generate_budget_pdf(budget, entity, months) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        Image as RLImage,
    )
    from app.services.pdf_generator import _load_logo, _FONT_NORMAL, _FONT_BOLD

    black    = colors.HexColor("#111111")
    gray_mid = colors.HexColor("#6b7280")
    gray_lt  = colors.HexColor("#f3f4f6")
    white    = colors.white
    hdr_bg   = colors.HexColor("#4B5563")
    pos_col  = colors.HexColor("#15803d")
    neg_col  = colors.HexColor("#dc2626")
    border_c = colors.HexColor("#e5e7eb")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=16*mm, leftMargin=16*mm,
                            topMargin=10*mm, bottomMargin=14*mm)
    PAGE_W = 178  # usable width in mm

    def sty(name, **kw):
        d = dict(fontName=_FONT_NORMAL, fontSize=9, textColor=black, leading=12)
        d.update(kw)
        return ParagraphStyle(name, **d)

    s_title  = sty("title", fontSize=13, fontName=_FONT_BOLD, leading=17)
    s_sub    = sty("sub",   fontSize=9,  textColor=gray_mid,  leading=13)
    s_locked = sty("lock",  fontSize=9,  fontName=_FONT_BOLD, textColor=neg_col, leading=13)
    s_hdr    = sty("hdr",   fontSize=7.5, fontName=_FONT_BOLD, textColor=white)
    s_hdr_r  = sty("hdrr",  fontSize=7.5, fontName=_FONT_BOLD, textColor=white, alignment=TA_RIGHT)
    s_lbl    = sty("lbl",   fontSize=8)
    s_note   = sty("note",  fontSize=7.5, textColor=gray_mid)
    s_val    = sty("val",   fontSize=8,  alignment=TA_RIGHT)
    s_bold   = sty("bold",  fontSize=8,  fontName=_FONT_BOLD)
    s_bold_r = sty("br",    fontSize=8,  fontName=_FONT_BOLD, alignment=TA_RIGHT)
    s_tot    = sty("tot",   fontSize=8,  fontName=_FONT_BOLD, textColor=white)
    s_tot_r  = sty("totr",  fontSize=8,  fontName=_FONT_BOLD, textColor=white, alignment=TA_RIGHT)
    # Tight variants for expense grids (6 value columns) — the amounts are one
    # unbreakable token each (NBSP), so they must FIT rather than wrap.
    s_val_sm   = sty("valsm", fontSize=7, alignment=TA_RIGHT)
    s_tot_r_sm = sty("totrsm", fontSize=7, fontName=_FONT_BOLD, textColor=white, alignment=TA_RIGHT)
    s_sec    = sty("sec",   fontSize=10, fontName=_FONT_BOLD, spaceBefore=4)

    def P(txt, style):
        return Paragraph(str(txt), style)

    story = []

    # ── Letterhead / logo (same fallbacks as the costing export) ──────────────
    lh_url  = getattr(entity, "letterhead_url",  None)
    lh_path = getattr(entity, "letterhead_path", None)
    added_header = False
    if lh_url or lh_path:
        lh_bytes = _load_logo(lh_url, lh_path)
        if lh_bytes:
            try:
                img = RLImage(io.BytesIO(lh_bytes), width=174*mm, height=105*mm, kind="proportional")
                img.hAlign = "CENTER"
                story.append(img)
                story.append(Spacer(1, 4*mm))
                added_header = True
            except Exception:
                pass
    if not added_header:
        logo_url  = getattr(entity, "logo_url",  None)
        logo_path = getattr(entity, "logo_path", None)
        if logo_url or logo_path:
            lb = _load_logo(logo_url, logo_path)
            if lb:
                try:
                    img = RLImage(io.BytesIO(lb), width=55*mm, height=30*mm, kind="proportional")
                    img.hAlign = "LEFT"
                    story.append(img)
                    story.append(Spacer(1, 3*mm))
                    added_header = True
                except Exception:
                    pass
        if not added_header:
            story.append(Paragraph(getattr(entity, "name", "") or "",
                                   sty("en", fontSize=14, fontName=_FONT_BOLD)))
            story.append(Spacer(1, 3*mm))

    # ── Title ─────────────────────────────────────────────────────────────────
    code = (getattr(entity, "code", "") or "").upper()
    stats = compute_budget_stats(budget, months, code)
    window_label = ", ".join(f"{MONTHS_SHORT[m]} {y}" for (m, y) in months)
    story.append(Paragraph(f"Budget — {MONTH_NAMES[budget.period_month]} {budget.period_year}", s_title))
    story.append(Paragraph(f"{code} — {getattr(entity, 'name', '') or ''} · Window: {window_label}", s_sub))
    if budget.locked_at is not None:
        locked_on = str(budget.locked_at)[:10]
        by = f" by {budget.locked_by_name}" if budget.locked_by_name else ""
        story.append(Paragraph(f"LOCKED{by} on {locked_on} — no values can be added or changed", s_locked))
    story.append(Spacer(1, 5*mm))

    # ── Summary ───────────────────────────────────────────────────────────────
    def kv_table(rows):
        data = [[P(lbl, s_bold if strong else s_lbl),
                 P(_fmt(val) if val is not None else "—",
                   sty(f"kv{idx}", fontSize=8, fontName=_FONT_BOLD, alignment=TA_RIGHT,
                       textColor=(pos_col if val is not None and val >= 0 else neg_col) if colored
                       else black))]
                for idx, (lbl, val, strong, colored) in enumerate(rows)]
        t = Table(data, colWidths=[90*mm, 60*mm], hAlign="LEFT")
        t.setStyle(TableStyle([
            ("GRID",          (0, 0), (-1, -1), 0.4, border_c),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [white, gray_lt]),
        ]))
        return t

    if stats["is_sft"]:
        b = stats["base"]
        rows = [
            ("Income", b["income"], False, False),
            ("Expenses (every expense section, To Pay + Paid)", b["expenses"], False, False),
            ("Profit (Income − Expenses)", b["profit"], True, True),
        ]
        if budget.external_profit is not None:
            rows.append(("Profit — Johan's Profit Sheet", _num(budget.external_profit), False, False))
        rows += [
            ("Personal Expenses (not on truck costing)", b["personal"], False, False),
            ("Actual Profit (Profit − Personal Expenses)", b["actual_profit"], True, True),
        ]
        story.append(Paragraph(f"Summary — {stats['base_label']} only", s_sec))
        story.append(Spacer(1, 1*mm))
        story.append(kv_table(rows))
        story.append(Paragraph(
            "The two months that follow are shown in the grid for cash-flow planning "
            "and don't affect these figures.", s_note))
    elif stats["month_split"]:
        g1, g2 = stats["month_split"]
        data = [
            [P("", s_hdr), P(g1["title"], s_hdr_r), P(g2["title"], s_hdr_r)],
            [P("Income", s_lbl), P(_fmt(g1["income"]), s_val), P(_fmt(g2["income"]), s_val)],
            [P("Expenses", s_lbl), P(_fmt(g1["expenses"]), s_val), P(_fmt(g2["expenses"]), s_val)],
            [P("Left Over", s_bold),
             P(_fmt(g1["left_over"]), sty("lo1", fontSize=8, fontName=_FONT_BOLD, alignment=TA_RIGHT,
                                          textColor=pos_col if g1["left_over"] >= 0 else neg_col)),
             P(_fmt(g2["left_over"]), sty("lo2", fontSize=8, fontName=_FONT_BOLD, alignment=TA_RIGHT,
                                          textColor=pos_col if g2["left_over"] >= 0 else neg_col))],
        ]
        t = Table(data, colWidths=[70*mm, 45*mm, 45*mm], hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), hdr_bg),
            ("GRID",          (0, 0), (-1, -1), 0.4, border_c),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(Paragraph("Summary", s_sec))
        story.append(Spacer(1, 1*mm))
        story.append(t)
        story.append(Paragraph(f"{g1['sub']} · {g2['sub']}", s_note))
    else:
        t = stats["totals"]
        story.append(Paragraph("Summary", s_sec))
        story.append(Spacer(1, 1*mm))
        story.append(kv_table([
            ("Income", t["income"], False, False),
            ("Expenses — To Pay", t["expenses_due"], False, False),
            ("Expenses — Paid", t["expenses_paid"], False, False),
            ("Left Over", t["left_over"], True, True),
        ]))
    story.append(Spacer(1, 4*mm))

    # ── Bank Info Summary ─────────────────────────────────────────────────────
    bank = stats["bank"]
    if bank:
        story.append(Paragraph("Bank Info Summary", s_sec))
        story.append(Spacer(1, 1*mm))
        rows = [[P(r.label or "", s_lbl), P(r.note or "", s_note), P(_fmt(r.amount) if r.amount is not None else "—", s_val)]
                for r in bank["accounts"]]
        cf_label = (bank["cash_flow_row"].label if bank["cash_flow_row"] is not None else "CASH FLOW TOTAL")
        rows.append([P(cf_label, s_bold),
                     P("" if (bank["cash_flow_row"] is not None and bank["cash_flow_row"].amount is not None)
                       else "Sum of the accounts above", s_note),
                     P(_fmt(bank["cash_flow_total"]), s_bold_r)])
        t = Table(rows, colWidths=[62*mm, 66*mm, 50*mm], hAlign="LEFT")
        t.setStyle(TableStyle([
            ("GRID",          (0, 0), (-1, -1), 0.4, border_c),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("ROWBACKGROUNDS", (0, 0), (-1, -2), [white, gray_lt]),
            ("BACKGROUND",    (0, len(rows) - 1), (-1, len(rows) - 1), gray_lt),
        ]))
        story.append(t)

        if stats["is_sft"]:
            story.append(Spacer(1, 2*mm))
            story.append(kv_table([
                (f"PROFIT FOR {stats['base_label'].upper()}", bank["profit"], True, True),
                ("vat back trailer", bank["vat_back"], False, False),
                ("profit would have been without vat back", bank["profit_excl_vat_back"], True, True),
            ]))
            if bank["to_be_paid"]:
                story.append(Spacer(1, 2*mm))
                story.append(Paragraph("TO BE PAID:", s_bold))
                tp_rows = [[P(r.label or "", s_lbl), P(_fmt(r.amount) if r.amount is not None else "—", s_val)]
                           for r in bank["to_be_paid"]]
                tp_rows.append([P("Total", s_bold), P(_fmt(bank["to_be_paid_total"]), s_bold_r)])
                t = Table(tp_rows, colWidths=[90*mm, 60*mm], hAlign="LEFT")
                t.setStyle(TableStyle([
                    ("GRID",          (0, 0), (-1, -1), 0.4, border_c),
                    ("LEFTPADDING",   (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
                    ("TOPPADDING",    (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("BACKGROUND",    (0, len(tp_rows) - 1), (-1, len(tp_rows) - 1), gray_lt),
                ]))
                story.append(t)
        story.append(Spacer(1, 4*mm))

    # ── Sections ──────────────────────────────────────────────────────────────
    for section in sorted(budget.sections, key=lambda s: (s.sort_order or 0, s.id)):
        is_income = section.section_type == "income"
        n_vals = len(months) * (1 if is_income else 2)
        val_w = min(30, (PAGE_W - 40) / n_vals)
        label_w = PAGE_W - val_w * n_vals
        col_w = [label_w*mm] + [val_w*mm] * n_vals
        tight = val_w < 24
        sv, stt = (s_val_sm, s_tot_r_sm) if tight else (s_val, s_tot_r)

        header = [P("Item", s_hdr)]
        for (m, y) in months:
            if is_income:
                header.append(P(f"{MONTHS_SHORT[m]} {y}", s_hdr_r))
            else:
                header.append(P(f"{MONTHS_SHORT[m]} To Pay", s_hdr_r))
                header.append(P(f"{MONTHS_SHORT[m]} Paid", s_hdr_r))

        rows = [header]
        tot = {(m, y, f): 0.0 for (m, y) in months for f in ("due", "paid")}
        for line in _sorted_lines(section):
            vmap = _value_map(line)
            row = [P(line.name or "", s_lbl)]
            for (m, y) in months:
                v = vmap.get((m, y))
                due = v.amount_due if v else None
                paid = v.amount_paid if v else None
                if due is not None:
                    tot[(m, y, "due")] += _num(due)
                if paid is not None:
                    tot[(m, y, "paid")] += _num(paid)
                row.append(P(_fmt(due) if due is not None else "", sv))
                if not is_income:
                    row.append(P(_fmt(paid) if paid is not None else "", sv))
            rows.append(row)

        if len(rows) > 1:
            tot_row = [P("Total", s_tot)]
            for (m, y) in months:
                tot_row.append(P(_fmt(tot[(m, y, "due")]), stt))
                if not is_income:
                    tot_row.append(P(_fmt(tot[(m, y, "paid")]), stt))
            rows.append(tot_row)

        story.append(Paragraph(section.name or "", s_sec))
        story.append(Spacer(1, 1*mm))
        if len(rows) == 1:
            story.append(Paragraph("No lines in this section.", s_note))
            story.append(Spacer(1, 3*mm))
            continue
        tbl = Table(rows, colWidths=col_w, repeatRows=1)
        style = [
            ("BACKGROUND",    (0, 0), (-1, 0), hdr_bg),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("GRID",          (0, 0), (-1, -1), 0.4, border_c),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND",    (0, len(rows) - 1), (-1, len(rows) - 1), colors.HexColor("#374151")),
        ]
        for i in range(1, len(rows) - 1):
            if i % 2 == 0:
                style.append(("BACKGROUND", (0, i), (-1, i), gray_lt))
        tbl.setStyle(TableStyle(style))
        story.append(tbl)
        story.append(Spacer(1, 4*mm))

    doc.build(story)
    return buf.getvalue()


# ── Excel ─────────────────────────────────────────────────────────────────────

def generate_budget_excel(budget, entity, months) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.services.pdf_generator import _load_logo

    code = (getattr(entity, "code", "") or "").upper()
    stats = compute_budget_stats(budget, months, code)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{MONTH_NAMES[budget.period_month]} {budget.period_year}"

    def fill(hex_str):
        return PatternFill("solid", fgColor=hex_str.lstrip("#"))

    def font(bold=False, color="111111", size=10):
        return Font(bold=bold, color=color, size=size)

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin = Side(style="thin", color="D1D5DB")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    NUM_FMT = "#,##0.00"

    max_vals = max(len(months) * 2, 3)
    ws.column_dimensions["A"].width = 44
    for i in range(2, max_vals + 2):
        ws.column_dimensions[get_column_letter(i)].width = 15

    row = 1

    # ── Letterhead image (same approach as the costing export) ────────────────
    from openpyxl.drawing.image import Image as XLImage
    lh_url  = getattr(entity, "letterhead_url",  None)
    lh_path = getattr(entity, "letterhead_path", None)
    lh_added = False
    if lh_url or lh_path:
        lh_bytes = _load_logo(lh_url, lh_path)
        if lh_bytes:
            try:
                from PIL import Image as PILImage
                pil_img = PILImage.open(io.BytesIO(lh_bytes))
                orig_w, orig_h = pil_img.size
                target_w_px = 600
                scale = target_w_px / orig_w
                target_h_px = int(orig_h * scale)
                out = io.BytesIO()
                pil_img.save(out, format="PNG")
                out.seek(0)
                xl_img = XLImage(out)
                xl_img.width  = target_w_px
                xl_img.height = target_h_px
                ws.add_image(xl_img, "A1")
                img_rows = max(1, target_h_px // 15)
                for _ in range(img_rows):
                    ws.append([])
                row += img_rows
                lh_added = True
            except Exception:
                pass
    if not lh_added:
        ws.cell(row=row, column=1, value=getattr(entity, "name", "") or "").font = font(bold=True, size=14)
        ws.row_dimensions[row].height = 22
        row += 1

    # ── Title ─────────────────────────────────────────────────────────────────
    window_label = ", ".join(f"{MONTHS_SHORT[m]} {y}" for (m, y) in months)
    ws.cell(row=row, column=1, value=f"Budget — {MONTH_NAMES[budget.period_month]} {budget.period_year}").font = \
        font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1,
            value=f"{code} — {getattr(entity, 'name', '') or ''} · Window: {window_label}").font = \
        font(color="6B7280", size=10)
    row += 1
    if budget.locked_at is not None:
        by = f" by {budget.locked_by_name}" if budget.locked_by_name else ""
        ws.cell(row=row, column=1,
                value=f"LOCKED{by} on {str(budget.locked_at)[:10]} — no values can be added or changed").font = \
            font(bold=True, color="DC2626", size=10)
        row += 1
    row += 1

    HDR_FILL   = fill("4B5563")
    ALT_FILL   = fill("F9FAFB")
    TOTAL_FILL = fill("374151")
    SEC_FILL   = fill("E5E7EB")
    HDR_FONT   = font(bold=True, color="FFFFFF", size=9)
    LBL_FONT   = font(size=9)
    TOT_WHITE  = font(bold=True, color="FFFFFF", size=9)
    NOTE_FONT  = font(color="6B7280", size=8)

    def amount_cell(r, c, value, fnt=None, fll=None, formula=False):
        cell = ws.cell(row=r, column=c, value=value)
        cell.font = fnt or LBL_FONT
        cell.alignment = align(h="right")
        cell.border = border_all
        cell.number_format = NUM_FMT
        if fll:
            cell.fill = fll
        return cell

    def label_cell(r, c, value, fnt=None, fll=None):
        cell = ws.cell(row=r, column=c, value=value)
        cell.font = fnt or LBL_FONT
        cell.alignment = align(wrap=True)
        cell.border = border_all
        if fll:
            cell.fill = fll
        return cell

    def kv_block(title, pairs):
        """Label/amount rows — the summary and bank-profit blocks."""
        nonlocal row
        if title:
            ws.cell(row=row, column=1, value=title).font = font(bold=True, size=11)
            row += 1
        for lbl, val, strong in pairs:
            label_cell(row, 1, lbl, fnt=font(bold=strong, size=9))
            amount_cell(row, 2, round(val, 2) if val is not None else None,
                        fnt=font(bold=strong, size=9))
            row += 1

    # ── Summary ───────────────────────────────────────────────────────────────
    if stats["is_sft"]:
        b = stats["base"]
        pairs = [
            ("Income", b["income"], False),
            ("Expenses (To Pay + Paid)", b["expenses"], False),
            ("Profit (Income − Expenses)", b["profit"], True),
        ]
        if budget.external_profit is not None:
            pairs.append(("Profit — Johan's Profit Sheet", _num(budget.external_profit), False))
        pairs += [
            ("Personal Expenses (not on truck costing)", b["personal"], False),
            ("Actual Profit (Profit − Personal Expenses)", b["actual_profit"], True),
        ]
        kv_block(f"Summary — {stats['base_label']} only", pairs)
    elif stats["month_split"]:
        g1, g2 = stats["month_split"]
        ws.cell(row=row, column=1, value="Summary").font = font(bold=True, size=11)
        row += 1
        label_cell(row, 1, "", fnt=HDR_FONT, fll=HDR_FILL)
        for c, g in ((2, g1), (3, g2)):
            cell = ws.cell(row=row, column=c, value=g["title"])
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = align(h="right")
            cell.border = border_all
        row += 1
        for lbl, key, strong in (("Income", "income", False), ("Expenses", "expenses", False),
                                 ("Left Over", "left_over", True)):
            label_cell(row, 1, lbl, fnt=font(bold=strong, size=9))
            for c, g in ((2, g1), (3, g2)):
                amount_cell(row, c, round(g[key], 2), fnt=font(bold=strong, size=9))
            row += 1
        ws.cell(row=row, column=1, value=f"{g1['sub']} · {g2['sub']}").font = NOTE_FONT
        row += 1
    else:
        t = stats["totals"]
        kv_block("Summary", [
            ("Income", t["income"], False),
            ("Expenses — To Pay", t["expenses_due"], False),
            ("Expenses — Paid", t["expenses_paid"], False),
            ("Left Over", t["left_over"], True),
        ])
    row += 1

    # ── Bank Info Summary ─────────────────────────────────────────────────────
    bank = stats["bank"]
    if bank:
        ws.cell(row=row, column=1, value="Bank Info Summary").font = font(bold=True, size=11)
        row += 1
        for r in bank["accounts"]:
            label_cell(row, 1, r.label or "")
            note = ws.cell(row=row, column=3, value=r.note or None)
            note.font = NOTE_FONT
            amount_cell(row, 2, round(_num(r.amount), 2) if r.amount is not None else None)
            row += 1
        cf_label = bank["cash_flow_row"].label if bank["cash_flow_row"] is not None else "CASH FLOW TOTAL"
        label_cell(row, 1, cf_label, fnt=font(bold=True, size=9), fll=ALT_FILL)
        amount_cell(row, 2, round(bank["cash_flow_total"], 2), fnt=font(bold=True, size=9), fll=ALT_FILL)
        row += 1

        if stats["is_sft"]:
            row += 1
            kv_block(None, [
                (f"PROFIT FOR {stats['base_label'].upper()}", bank["profit"], True),
                ("vat back trailer", bank["vat_back"], False),
                ("profit would have been without vat back", bank["profit_excl_vat_back"], True),
            ])
            if bank["to_be_paid"]:
                row += 1
                ws.cell(row=row, column=1, value="TO BE PAID:").font = font(bold=True, size=10)
                row += 1
                for r in bank["to_be_paid"]:
                    label_cell(row, 1, r.label or "")
                    amount_cell(row, 2, round(_num(r.amount), 2) if r.amount is not None else None)
                    row += 1
                label_cell(row, 1, "Total", fnt=font(bold=True, size=9), fll=ALT_FILL)
                amount_cell(row, 2, round(bank["to_be_paid_total"], 2), fnt=font(bold=True, size=9), fll=ALT_FILL)
                row += 1
        row += 1

    # ── Sections ──────────────────────────────────────────────────────────────
    for section in sorted(budget.sections, key=lambda s: (s.sort_order or 0, s.id)):
        is_income = section.section_type == "income"
        n_vals = len(months) * (1 if is_income else 2)

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + n_vals)
        c = ws.cell(row=row, column=1, value=section.name or "")
        c.font = font(bold=True, size=11)
        c.fill = SEC_FILL
        c.alignment = align()
        ws.row_dimensions[row].height = 18
        row += 1

        label_cell(row, 1, "Item", fnt=HDR_FONT, fll=HDR_FILL)
        col = 2
        for (m, y) in months:
            labels = [f"{MONTHS_SHORT[m]} {y}"] if is_income else [f"{MONTHS_SHORT[m]} To Pay", f"{MONTHS_SHORT[m]} Paid"]
            for lbl in labels:
                cell = ws.cell(row=row, column=col, value=lbl)
                cell.font = HDR_FONT
                cell.fill = HDR_FILL
                cell.alignment = align(h="right")
                cell.border = border_all
                col += 1
        row += 1

        lines = _sorted_lines(section)
        first_line_row = row
        for i, line in enumerate(lines):
            vmap = _value_map(line)
            fll = ALT_FILL if i % 2 else None
            label_cell(row, 1, line.name or "", fll=fll)
            col = 2
            for (m, y) in months:
                v = vmap.get((m, y))
                fields = ("amount_due",) if is_income else ("amount_due", "amount_paid")
                for f in fields:
                    val = getattr(v, f, None) if v else None
                    amount_cell(row, col, round(_num(val), 2) if val is not None else None, fll=fll)
                    col += 1
            row += 1

        if lines:
            # Totals as SUM formulas over the line rows, so the sheet stays live.
            label_cell(row, 1, "Total", fnt=TOT_WHITE, fll=TOTAL_FILL)
            for col in range(2, 2 + n_vals):
                letter = get_column_letter(col)
                amount_cell(row, col, f"=SUM({letter}{first_line_row}:{letter}{row - 1})",
                            fnt=TOT_WHITE, fll=TOTAL_FILL)
            row += 1
        else:
            ws.cell(row=row, column=1, value="No lines in this section.").font = NOTE_FONT
            row += 1
        row += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
